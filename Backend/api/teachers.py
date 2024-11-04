# Django
from django.db import IntegrityError, transaction
from django.core.files.storage import default_storage


# Document Manipulation
import pandas as pd
from openpyxl.styles import Font, Alignment, Protection
from openpyxl import Workbook, load_workbook

# Django Restframework
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from api.models import *
from api.serializer import *
from api.utils import *
import json
from collections import defaultdict
from datetime import datetime
import hashlib
import math
from decimal import Decimal

# TEACHER
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_teacher_data(request):
    teacher = request.user.staff
    school = teacher.school
    current_term = int(request.GET.get('term'))
    current_academic_year = AcademicYear.objects.get(school=school, name=request.GET.get('year'))
    department_data = None
    subject_assignments = SubjectAssignment.objects.prefetch_related('subjects', 'students_class__students').filter(school=school, teacher=teacher, academic_year=current_academic_year, academic_term=current_term)
    subject_assignments_data = TeacherSubjectAssignmentSerializer(subject_assignments, many=True).data
    
    department = Department.objects.filter(school=school, teachers=teacher).first()
    if department:
        department_data = DepartmentNameSubjectsTeachersDataSerializer(department).data
    
    return Response({
        'subject_assignments': subject_assignments_data,
        'department_data': department_data,
    })
    
    
@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def teacher_students_attendance(request):
    teacher = request.user.staff
    school = teacher.school
    
    if request.method == 'GET':
        head_classes = Classe.objects.prefetch_related('students').filter(school=school, head_teacher=teacher).order_by('name')
        all_attendances = []
        current_academic_year = AcademicYear.objects.get(school=school, name=request.GET.get('year'))
        if head_classes.exists():
            for _class in head_classes:
                data = {'class_name': _class.name, 'students': [{'name': f"{st.user.first_name} {st.user.last_name}", 'st_id': st.st_id} for st in _class.students.all()]}
                attendance = StudentAttendance.objects.filter(school=school, students_class=_class, academic_year=current_academic_year).order_by('-date')
                attendance_data = StudentsAttendanceSerializer(attendance, many=True).data
                data['attendances'] = attendance_data
                all_attendances.append(data)
            
        return Response({
            'students_attendance': all_attendances,
        })
    
    if request.method == 'POST':
        data = request.data
        students_class = Classe.objects.prefetch_related('students').get(school=school, name=data['className'])
        academic_year = AcademicYear.objects.get(school=school, name=data['year'])
        if data['type'] == 'create':
            students = data['absentStudents'].split(',')
            try:
                existing_attendance = StudentAttendance.objects.get(
                    school=teacher.school,
                    date=data['date'],
                    teacher=teacher,
                    students_class=students_class,
                    academic_year=academic_year,
                    academic_term=int(data['term']),
                )
                return Response({
                    'message': f"You have already uploaded attendance for this date [ {data['date']} ]",
                }, status=400)

            except StudentAttendance.DoesNotExist:
                with transaction.atomic():
                    attendance = StudentAttendance.objects.create(
                        school=teacher.school,
                        teacher=teacher,
                        date=data['date'],
                        students_class=students_class,
                        academic_year=academic_year,
                        academic_term=int(data['term']),
                    )
                        
                    for st in students_class.students.all():
                        if st.st_id not in students:
                            attendance.students_present.add(st)
                        else:
                            attendance.students_absent.add(st)

                    attendance.save()

                new_attendance = StudentsAttendanceSerializer(StudentAttendance.objects.get(
                    school=teacher.school,
                    academic_year=academic_year,
                    academic_term=int(data['term']),
                    date=data['date'],
                    students_class=students_class,
                    teacher=teacher
                )).data

                return Response(new_attendance)

        elif data['type'] == 'delete':
            try:
                students_class = Classe.objects.get(school=teacher.school, name=data['className'])
                with transaction.atomic():
                    student_attendance = StudentAttendance.objects.get(
                        school=teacher.school,
                        academic_year=academic_year,
                        academic_term=int(data['term']),
                        students_class=students_class,
                        teacher=teacher,
                        date=data['date'],
                    )
                    student_attendance.delete()

                return Response()

            except StudentAttendance.DoesNotExist:
                return Response(status=400)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def teacher_assessments(request):
    teacher = request.user.staff
    school = teacher.school
    
    if request.method == 'GET':
        current_term = int(request.GET.get('term'))
        current_academic_year = AcademicYear.objects.get(school=school, name=request.GET.get('year'))
        all_assessments_data = []
            
        subject_assignments = list(SubjectAssignment.objects.select_related('teacher', 'students_class').prefetch_related('subjects').filter(school=school, teacher=teacher, academic_year=current_academic_year, academic_term=current_term))
        for assign in subject_assignments:
            all_students = assign.students_class.students.all()
            assessment_class_data = {'class_name': assign.students_class.name, 'assignments': []}
            for _subject in assign.subjects.all():
                assessment_subject_data = {'subject': _subject.name, 'assessments': []}
                all_assessments = list(Assessment.objects.select_related('student').filter(school=school, teacher=teacher, subject=_subject, student_class=assign.students_class, academic_year=current_academic_year, academic_term=current_term))
                assessment_titles = list(set([x.title for x in all_assessments]))
                for _title in assessment_titles:
                    assessment_with_title = Assessment.objects.select_related('student').filter(school=school, teacher=teacher, subject=_subject, title=_title, student_class=assign.students_class, academic_year=current_academic_year, academic_term=current_term).order_by('-score')
                    first_assessment = assessment_with_title[0]
                    assessment_title_data = {'title': _title, 'total_score': first_assessment.total_score, 'description': first_assessment.description, 'percentage': first_assessment.percentage, 'assessment_date': first_assessment.assessment_date, 'students_with_assessment': [], 'students_without_assessment': []}
                    students_with_assessments = set()
                    for _assessment in assessment_with_title:
                        if _assessment.student:
                            students_with_assessments.add(_assessment.student)
                            assessment_title_data['students_with_assessment'].append({
                                'name': f"{_assessment.student.user.first_name} {_assessment.student.user.last_name}",
                                'st_id': _assessment.student.st_id,
                                'score': _assessment.score,
                                'comment': _assessment.comment,
                            })
                    for _student in all_students:
                        if _student not in students_with_assessments:
                            assessment_title_data['students_without_assessment'].append({
                                'name': f"{_student.user.first_name} {_student.user.last_name}",
                                'st_id': _student.st_id,
                            })
                    assessment_subject_data['assessments'].append(assessment_title_data)
                assessment_class_data['assignments'].append(assessment_subject_data)
            all_assessments_data.append(assessment_class_data)
            
        return Response({
            'assessments': all_assessments_data
        })
    
    elif request.method == 'POST':
        year = request.data['year']
        term = int(request.data['term'])
        subject_name = request.data['subject']
        st_class_name = request.data['studentsClassName']
        current_academic_year = AcademicYear.objects.select_related('period_division').get(school=school, name=year)
        assessment_title = request.data['title']
        if request.data['type'] == 'getFile':
            students = json.loads(request.data['selectedStudents'])
            total_score = request.data['totalScore']
            if school.students_id:
                data = [
                    ['STUDENT NAME', 'STUDENT ID', 'SCORE', 'COMMENT']
                ]
            else:
                data = [
                    ['STUDENT NAME', 'STUDENT USERNAME', 'SCORE', 'COMMENT']
                ]

            if len(students) == 0:
                return Response({
                    'message': f"You have already uploaded all the {subject_name}[{assessment_title}] Assessment scores for all students in this class",
                }, status=400)
                
            subject_obj = Subject.objects.get(name=subject_name)
            for _st in students:
                row = [_st['name'], _st['st_id'], '']
                data.append(row)

            wb = Workbook()
            ws = wb.active
            ws.title = subject_name

            ws.merge_cells('A1:I3')
            ws['A1'].value = f"SUBJECT: [{subject_name}]  ACADEMIC YEAR: [{year}]  CLASS: [{st_class_name}]  {current_academic_year.period_division.name}: [{term}]  ASSESSMENT: [{assessment_title}]"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            count = 0
            container = ws['B5:E200']
            while count < len(data):
                for row_ws, row_data in zip(container, data):
                    for cell, item in zip(row_ws, row_data):
                        cell.value = item

                count += 1

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 50

            for row in ws['A3:F200']:
                for cell in row:
                    cell.font = Font(size=14)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in ws['B3:B200']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            for cells in ws['A5:F200'][0]:
                cells.font = Font(bold=True, size=14)

            ws.protection.sheet = True
            for cell in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=4, max_col=4):
                cell[0].protection = Protection(locked=False)
            for cell in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=5, max_col=5):
                cell[0].protection = Protection(locked=False)

            ws.protection.password = 'teamjn'

            filename = f"{subject_name.replace(' ', '-')}-{st_class_name}-{assessment_title}.xlsx"
            byte_file = io.BytesIO()
            wb.save(byte_file)
            file_path = f"{get_school_folder(school.name)}/staff/{teacher.user.username}/{filename}"
            if default_storage.exists(file_path):
                default_storage.delete(file_path)

            saved_file = default_storage.save(file_path, byte_file)
            if settings.DEBUG:
                return Response({
                    'filename': filename,
                    'message': 'File generated successfully',
                    'file_path': f"http://localhost:8000{default_storage.url(saved_file)}",
                })
            else:
                return Response({
                    'filename': filename,
                    'message': 'File generated successfully',
                    'file_path': default_storage.url(saved_file),
                })

        elif request.data['type'] == 'createAssessment':
            students_class = Classe.objects.get(school=school, name=st_class_name)
            subject_obj = Subject.objects.get(name=subject_name)
            description = request.data['description']
            total_score = float(request.data['totalScore'])
            assessment_date = request.data['date']
            assessment_date_object = datetime.strptime(assessment_date, '%Y-%m-%d').date()
            existing_results = StudentResult.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).exists()
            if existing_results:
                return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
            if total_score <= 0:
                return Response({'message': f"The total score of the assessment cannot be negative or zero(0)"}, status=400)
            if (assessment_date_object > current_academic_year.end_date) or (assessment_date_object < current_academic_year.start_date):
                return Response({'message': f"The assessment date must be between the academic year start and end dates. That's {current_academic_year.start_date} and {current_academic_year.end_date}"}, status=400)
            if Assessment.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term).exists():
                return Response({'message': f"An assessment with the title '{assessment_title}' already exists. Please use a different title."}, status=400)
            
            with transaction.atomic():
                assessment_obj = Assessment.objects.create(
                    school=school, 
                    teacher=teacher, 
                    subject=subject_obj, 
                    student_class=students_class,
                    title=assessment_title,
                    description=description,
                    total_score=float(total_score),
                    assessment_date=assessment_date,
                    academic_year=current_academic_year, 
                    academic_term=term
                )
                try:
                    assessment_obj.save()
                except Exception:
                    return Response(status=400) 
            
            return Response({'message': "Operation successful"})
            
        elif request.data['type'] == 'uploadWithFile':
            file = request.data['file']
            subject_obj = Subject.objects.get(name=subject_name)
            old_assessment = Assessment.objects.filter(school=school, teacher=teacher, subject=subject_obj, title=assessment_title, academic_year=current_academic_year, academic_term=term).first()
            description = old_assessment.description
            total_score = old_assessment.total_score
            percentage = old_assessment.percentage
            assessment_date = old_assessment.assessment_date
                
            try:
                wb = load_workbook(file)
            except Exception:
                return Response({'message': "Invalid file. Ensure you upload the excel file that you generated"}, status=400)
            
            ws = wb.active
            ws.protection = False
            if ws['A1'].value != f"SUBJECT: [{subject_name}]  ACADEMIC YEAR: [{year}]  CLASS: [{st_class_name}]  {current_academic_year.period_division.name}: [{term}]  ASSESSMENT: [{assessment_title}]":
                return Response({'message': "Invalid file. Ensure you upload the same excel file that you got"}, status=400)
            
            cleaned_data = []
            for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=5):
                data = []
                for cell in row:
                    data.append(cell.value)
                cleaned_data.append(data)
            
            if len(cleaned_data) == 0:
                return Response({'message': f"There are no students data in the uploaded file"}, status=400)

            valid_rows = [x for x in cleaned_data if x[0] and x[1]]
            for row in valid_rows:
                score = row[2]
                try:
                    float(score)
                except Exception:
                    return Response({'message': f"The scores must be a number. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)
                
                if score > total_score:
                    print(total_score)
                    return Response({'message': f"Scores cannot exceed the total score for the assessment. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)
                elif score < 0:
                    return Response({'message': f"Scores cannot be negative. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)

            students_class = Classe.objects.prefetch_related('students').get(school=school, name=st_class_name)
            assessment_students = []
            with transaction.atomic():
                assessments_to_create = []
                for _student in valid_rows:
                    if students_class.students.filter(school=school, st_id=_student[1]).exists():
                        student = Student.objects.get(school=school, st_id=_student[1])
                        assessment_students.append(student)
                        st_assessment = Assessment(
                            student=student,
                            subject=subject_obj,
                            teacher=teacher,
                            student_class=students_class,
                            score=float(_student[2]),
                            academic_year=current_academic_year,
                            title=assessment_title,
                            description=description,
                            percentage=percentage,
                            total_score=total_score,
                            comment=_student[3] if _student[3] else '',
                            academic_term=term,
                            school=school,
                            assessment_date=assessment_date,
                        )
                        assessments_to_create.append(st_assessment)
                    else:
                        transaction.set_rollback(True)
                        return Response({'message': f"Invalid student {_student[0]}[{_student[1]}]. Make sure you don't change anything in the excel file except the scores and comments"}, status=400)
                try:
                    Assessment.objects.bulk_create(assessments_to_create)
                    assessment_to_delete = Assessment.objects.get(school=school, teacher=teacher, student=None, student_class=students_class, subject=subject_obj, title=assessment_title, academic_year=current_academic_year, academic_term=term)
                    assessment_to_delete.delete()
                except Assessment.DoesNotExist:
                    transaction.set_rollback(False)
                except IntegrityError:
                    transaction.set_rollback(True)
                    return Response({'message': f"You have already uploaded Assessment for some of the students in the file. Click on get file to get a new updated file"}, status=400)
                except Exception:
                    transaction.set_rollback(True)
                    return Response({'message': f"An unexpected error occurred! Ensure you don't delete or change anything in the excel file except the scores and comments"}, status=400)
            
            assessments = Assessment.objects.select_related('student__user').filter(school=school, teacher=teacher, subject=subject_obj, title=assessment_title, student__in=assessment_students, academic_year=current_academic_year, academic_term=term)
            students_data = [{
                'name': f"{x.student.user.first_name} {x.student.user.last_name}", 
                'st_id': x.student.st_id,
                'score': x.score,
                'comment': x.comment,
            } for x in assessments]
            return Response({'message': "Operation successful!", 'data': students_data})

        elif request.data['type'] == 'uploadWithoutFile':
            students_class = Classe.objects.get(school=school, name=st_class_name)
            subject_obj = Subject.objects.get(name=subject_name)
            student_ids = json.loads(request.data['selectedStudents'])
            comment = request.data['comment']
            old_assessment = Assessment.objects.filter(school=school, teacher=teacher, student_class=students_class, subject=subject_obj, title=assessment_title, academic_year=current_academic_year, academic_term=term).first()
            description = old_assessment.description
            total_score = old_assessment.total_score
            percentage = old_assessment.percentage
            assessment_date = old_assessment.assessment_date
            
            score = float(request.data['score'])
            if score > total_score:
                return Response({'message': f"The student(s) score cannot exceed the total score for the assessment!"}, status=400)
            
            with transaction.atomic():
                assessments_to_create = []
                for _st_id in student_ids:
                    student = Student.objects.get(school=school, st_id=_st_id)
                    st_assessment = Assessment(
                        student=student,
                        subject=subject_obj,
                        teacher=teacher,
                        student_class=students_class,
                        score=score,
                        academic_year=current_academic_year,
                        title=assessment_title,
                        description=description,
                        percentage=percentage,
                        total_score=total_score,
                        comment=comment,
                        academic_term=term,
                        school=school,
                        assessment_date=assessment_date,
                    )
                    assessments_to_create.append(st_assessment)
                
                try:
                    Assessment.objects.bulk_create(assessments_to_create)
                    assessment_to_delete = Assessment.objects.get(school=school, teacher=teacher, student=None, student_class=students_class, subject=subject_obj, title=assessment_title, academic_year=current_academic_year, academic_term=term)
                    assessment_to_delete.delete()
                except Assessment.DoesNotExist:
                    transaction.set_rollback(False)
                except Exception:
                    transaction.set_rollback(True)
                    return Response(status=400)
            
            return Response({'message': "Operation successful!"})

        elif request.data['type'] == 'editAssessment':
            students_class = Classe.objects.get(school=school, name=st_class_name)
            subject_obj = Subject.objects.get(name=subject_name)
            if (request.data['editType'] == 'title'):
                with transaction.atomic():
                    assessments_to_update = []
                    old_title = assessment_title
                    new_title = request.data['newTitle']
                    if Assessment.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, title=new_title, academic_year=current_academic_year, academic_term=term).exists():
                        return Response({'message': f"Assessment with title [ {new_title} ] already exists! Use a different title"}, status=400)
                    assessments = list(Assessment.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, title=old_title, academic_year=current_academic_year, academic_term=term))
                    for _assessment in assessments:
                        _assessment.title = new_title
                        assessments_to_update.append(_assessment)
                    try:
                        Assessment.objects.bulk_update(assessments_to_update, ['title'])
                    except Exception:
                        return Response(status=400) 
                
            elif (request.data['editType'] == 'description'):
                with transaction.atomic():
                    assessments_to_update = []
                    new_description = request.data['newDescription']
                    assessments = list(Assessment.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term))
                    for _assessment in assessments:
                        _assessment.description = new_description
                        assessments_to_update.append(_assessment)
                    try:
                        Assessment.objects.bulk_update(assessments_to_update, ['description'])
                    except Exception:
                        return Response(status=400) 
                    
            elif (request.data['editType'] == 'totalScore'):
                assessments_to_update = []
                results_to_update = []
                new_total_score = Decimal(request.data['newTotalScore'])
                if new_total_score <= 0:
                    return Response({'message': "The total score of the assessment cannot be negative or zero(0)"}, status=400)
                assessments = Assessment.objects.select_related('student').filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term)
                grading_system = GradingSystem.objects.filter(schools=school).order_by('lower_limit')
                for _assessment in assessments:
                    old_total_score = _assessment.total_score
                    if _assessment.score and new_total_score < float(_assessment.score):
                        return Response({'message': "The total score of the assessment cannot be less than the a student's score"}, status=400)
                    _assessment.total_score = new_total_score
                    assessments_to_update.append(_assessment)
                    st_result = StudentResult.objects.filter(school=school, teacher=teacher, student=_assessment.student, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).first()
                    if st_result:
                        old_assessment_score = (_assessment.score/old_total_score)*_assessment.percentage
                        new_assessment_score = (_assessment.score/new_total_score)*_assessment.percentage
                        new_total_assessment_score = (st_result.total_assessment_score + new_assessment_score) - old_assessment_score
                        new_total_assessment_score = math.ceil(new_total_assessment_score) if new_total_assessment_score - math.floor(new_total_assessment_score) >= 0.5 else math.floor(new_total_assessment_score)
                        new_result = st_result.exam_score + new_total_assessment_score
                        student_grade = next((item for item in grading_system if item.upper_limit >= new_result >= item.lower_limit), None)
                        st_result.total_assessment_score = new_total_assessment_score
                        st_result.result = new_result
                        st_result.grade = student_grade.label
                        st_result.remark = student_grade.remark
                        results_to_update.append(st_result)
                with transaction.atomic():
                    try:
                        Assessment.objects.bulk_update(assessments_to_update, ['total_score'])
                        StudentResult.objects.bulk_update(results_to_update, ['total_assessment_score', 'result', 'grade', 'remark'])
                    except Exception:
                        return Response(status=400) 
                all_results =  StudentResult.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).order_by('-result')
                all_results_data = TeacherGetStudentResultSerializer(all_results, many=True).data
                return Response(all_results_data)
            
            elif (request.data['editType'] == 'date'):
                assessments_to_update = []
                new_date = request.data['newDate']
                new_date_object = datetime.strptime(new_date, '%Y-%m-%d').date()
                if (new_date_object > current_academic_year.end_date) or (new_date_object < current_academic_year.start_date):
                    return Response({'message': f"The assessment date must be between the academic year start and end dates. That's {current_academic_year.start_date} and {current_academic_year.end_date}"}, status=400)
                assessments = list(Assessment.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term))
                for _assessment in assessments:
                    _assessment.assessment_date = new_date
                    assessments_to_update.append(_assessment)
                with transaction.atomic():
                    try:
                        Assessment.objects.bulk_update(assessments_to_update, ['assessment_date'])
                    except Exception:
                        return Response({'message': 'An unexpected error occurred! try again later'}, status=400)
                
            elif (request.data['editType'] == 'comment'):
                new_comment = request.data['newComment']
                student = Student.objects.get(school=school, st_id=request.data['studentId'])
                with transaction.atomic():
                    try:
                        assessment = Assessment.objects.get(school=school, teacher=teacher, subject=subject_obj, student=student, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term)
                        assessment.comment = new_comment
                        assessment.save()
                    except Exception:
                        return Response(status=400)     
                
            elif (request.data['editType'] == 'score'):
                new_score = Decimal(request.data['newScore'])
                if new_score < 0:
                    return Response({'message': "The student's score cannot be negative"}, status=400)
                student = Student.objects.get(school=school, st_id=request.data['studentId'])
                with transaction.atomic():
                    try:
                        assessment = Assessment.objects.get(school=school, teacher=teacher, subject=subject_obj, student=student, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term)
                        if new_score > float(assessment.total_score):
                            return Response({"message": "The student's score cannot exceed the total assessment score"}, status=400)
                        old_assessment_score = (assessment.score/assessment.total_score)*assessment.percentage
                        assessment.score = new_score
                        assessment.save()
                        st_result = StudentResult.objects.get(school=school, teacher=teacher, subject=subject_obj, student=student, student_class=students_class, academic_year=current_academic_year, academic_term=term)
                        new_assessment_score = (new_score/assessment.total_score)*assessment.percentage
                        new_total_assessment_score = (st_result.total_assessment_score + new_assessment_score) - old_assessment_score
                        new_total_assessment_score = math.ceil(new_total_assessment_score) if new_total_assessment_score - math.floor(new_total_assessment_score) >= 0.5 else math.floor(new_total_assessment_score)
                        grading_system = GradingSystem.objects.filter(schools=school).order_by('lower_limit')
                        new_result = new_total_assessment_score + st_result.exam_score
                        student_grade = next((item for item in grading_system if item.upper_limit >= new_result >= item.lower_limit), None)
                        st_result.total_assessment_score = new_total_assessment_score
                        st_result.result = new_result
                        st_result.grade = student_grade.label
                        st_result.remark = student_grade.remark
                        st_result.save()
                        return Response({
                            'new_total_assessment_score': new_total_assessment_score, 
                            'new_result': new_result,
                            'new_grade': student_grade.label,
                            'new_remark': student_grade.remark,
                        })
                    except StudentResult.DoesNotExist:
                        transaction.set_rollback(False)
                    except Exception:
                        return Response(status=400)                
                
            return Response()
        
        elif request.data['type'] == 'deleteAssessment':
            students_class = Classe.objects.get(school=school, name=st_class_name)
            subject_obj = Subject.objects.get(name=subject_name)
            with transaction.atomic():
                try:
                    existing_results = StudentResult.objects.filter(school=school, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term).exists()
                    if existing_results:
                        return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
                    assessments = Assessment.objects.filter(school=school, teacher=teacher, student_class=students_class, subject=subject_obj, title=assessment_title, academic_year=current_academic_year, academic_term=term)
                    assessments.delete()
                except Exception:
                    transaction.set_rollback(True)
                    return Response(status=400)
            
            return Response()
    

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def teacher_exams(request):
    teacher = request.user.staff
    school = teacher.school
    
    if request.method == 'GET':
        current_term = int(request.GET.get('term'))
        current_academic_year = AcademicYear.objects.get(school=school, name=request.GET.get('year'))
        all_exams_data = []
        subject_assignments = SubjectAssignment.objects.select_related('teacher', 'students_class').prefetch_related('subjects').filter(school=school, teacher=teacher, academic_year=current_academic_year, academic_term=current_term)
        for assign in subject_assignments:
            all_students = assign.students_class.students.all()
            exams_data = {'class_name': assign.students_class.name, 'exams': []}
            for _subject in assign.subjects.all():
                subject_data = {'subject': _subject.name, 'students_with_exams': [], 'students_without_exams': []}
                all_exams = Exam.objects.select_related('student__user').filter(school=school, teacher=teacher, subject=_subject, student_class=assign.students_class, academic_year=current_academic_year, academic_term=current_term).order_by('-score')
                students_with_exams = set()
                if all_exams.exists():
                    subject_data['percentage'] = all_exams.first().percentage
                    subject_data['total_score'] = all_exams.first().total_score
                for _exams in all_exams:
                    if _exams.student:
                        subject_data['students_with_exams'].append({
                            'name': f"{_exams.student.user.first_name} {_exams.student.user.last_name}",
                            'st_id': _exams.student.st_id,
                            'score': _exams.score,
                        })
                        students_with_exams.add(_exams.student)
                for _student in all_students:
                    if _student not in students_with_exams:
                        subject_data['students_without_exams'].append({
                            'name': f"{_student.user.first_name} {_student.user.last_name}",
                            'st_id': _student.st_id,
                        })
                exams_data['exams'].append(subject_data)
            all_exams_data.append(exams_data)
            
        return Response({
            'exams_data': all_exams_data
        })
    
    elif request.method == 'POST':
        year = request.data['year']
        term = int(request.data['term'])
        subject_name = request.data['subject']
        st_class_name = request.data['studentsClassName']
        current_academic_year = AcademicYear.objects.select_related('period_division').get(school=school, name=year)
        if request.data['type'] == 'getFile':
            students = json.loads(request.data['selectedStudents'])
            if school.students_id:
                data = [
                    ['STUDENT NAME', 'STUDENT ID', 'SCORE']
                ]
            else:
                data = [
                    ['STUDENT NAME', 'STUDENT USERNAME', 'SCORE']
                ]

            if len(students) == 0:
                return Response({
                    'message': f"You have already uploaded all the {subject_name} Exams scores for all students in this class",
                }, status=400)
                
            for _st in students:
                row = [_st['name'], _st['st_id'], '']
                data.append(row)

            wb = Workbook()
            ws = wb.active
            ws.title = subject_name

            ws.merge_cells('A1:I3')
            ws['A1'].value = f"SUBJECT: [{subject_name}]  ACADEMIC YEAR: [{year}]  CLASS: [{st_class_name}]  {current_academic_year.period_division.name} {term} EXAMS"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            count = 0
            container = ws['B5:D200']
            while count < len(data):
                for row_ws, row_data in zip(container, data):
                    for cell, item in zip(row_ws, row_data):
                        cell.value = item

                count += 1

            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15

            for row in ws['A3:F200']:
                for cell in row:
                    cell.font = Font(size=14)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in ws['B3:B200']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            for cells in ws['A5:F200'][0]:
                cells.font = Font(bold=True, size=14)

            ws.protection.sheet = True
            for cell in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=4, max_col=4):
                cell[0].protection = Protection(locked=False)

            ws.protection.password = 'teamjn'

            filename = f"{subject_name.replace(' ', '-')}-{st_class_name}-Exams.xlsx"
            byte_file = io.BytesIO()
            wb.save(byte_file)
            file_path = f"{get_school_folder(school.name)}/staff/{teacher.user.username}/{filename}"
            if default_storage.exists(file_path):
                default_storage.delete(file_path)

            saved_file = default_storage.save(file_path, byte_file)
            if settings.DEBUG:
                return Response({
                    'filename': filename,
                    'message': 'File generated successfully',
                    'file_path': f"http://localhost:8000{default_storage.url(saved_file)}",
                })
            else:
                return Response({
                    'filename': filename,
                    'message': 'File generated successfully',
                    'file_path': default_storage.url(saved_file),
                })
        
        elif request.data['type'] == 'createExams':
            subject_obj = Subject.objects.get(name=subject_name)
            students_class = Classe.objects.get(school=school, name=st_class_name)
            existing_exams = Exam.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).exists()
            if existing_exams:
                return Response({'message': f"You have already created {subject_name} exams for this class"}, status=400)
            existing_results = StudentResult.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).exists()
            if existing_results:
                return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
            exams_total_score = request.data['totalScore']
            exams_obj = Exam.objects.create(
                school=school,
                teacher=teacher,
                student_class=students_class,
                subject=subject_obj,
                percentage=0,
                score=0,
                total_score=exams_total_score,
                academic_year=current_academic_year,
                academic_term=term
            )
            with transaction.atomic():
                try:
                    exams_obj.save()
                except Exception:
                    return Response(status=400)
                
            return Response({'message': "Operation Successful"})
            
        elif request.data['type'] == 'uploadWithFile':
            file = request.data['file']
            try:
                wb = load_workbook(file)
            except Exception:
                return Response({'message': "Invalid file. Ensure you upload the excel file that you generated"}, status=400)
            
            ws = wb.active
            ws.protection = False
            if ws['A1'].value != f"SUBJECT: [{subject_name}]  ACADEMIC YEAR: [{year}]  CLASS: [{st_class_name}]  {current_academic_year.period_division.name} {term} EXAMS":
                return Response({'message': "Invalid file. Ensure you upload the excel file that you generated"}, status=400)
            
            cleaned_data = []
            for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=4):
                data = []
                for _cell in row:
                    data.append(_cell.value)
                cleaned_data.append(data)
            
            valid_rows = [x for x in cleaned_data if x[0]]
            if len(valid_rows) == 0:
                return Response({'message': f"There are no students data in the uploaded file"}, status=400)
            for row in valid_rows:
                score = row[2]
                try:
                    float(score)
                except Exception:
                    return Response({'message': f"The scores must be a number. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)
                if score > 100:
                    return Response({'message': f"Scores cannot be more than 100. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)
                elif score < 0:
                    return Response({'message': f"Scores cannot be negative. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)

            students_class = Classe.objects.prefetch_related('students').get(school=school, name=st_class_name)
            exams_students = []
            subject_obj = Subject.objects.get(name=subject_name)
            exams_to_create = []
            for _student in valid_rows:
                if students_class.students.filter(school=school, st_id=_student[1]).exists():
                    st = Student.objects.get(school=school, st_id=_student[1])
                    exams_students.append(st)
                    st_exam = Exam(
                        student=st,
                        subject=subject_obj,
                        teacher=teacher,
                        percentage=0,
                        student_class=students_class,
                        score=float(_student[2]),
                        academic_year=current_academic_year,
                        academic_term=term,
                        school=school,
                    )
                    exams_to_create.append(st_exam)
                else:
                    return Response({'message': f"Invalid student {_student[0]}[{_student[1]}]. Make sure you don't change anything in the excel file except the scores"}, status=400)
            with transaction.atomic():
                try:
                    Exam.objects.bulk_create(exams_to_create)
                    exam_create_obj = Exam.objects.get(school=school, teacher=teacher, student=None, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term)
                    exam_create_obj.delete()
                except Exam.DoesNotExist:
                    transaction.set_rollback(False)
                except IntegrityError:
                    transaction.set_rollback(True)
                    return Response({'message': f"You have already uploaded exams score for some of the students in the file. Click on get file to get an updated file"}, status=400)
                except Exception:
                    transaction.set_rollback(True)
                    return Response({'message': f"Invalid data! Ensure you don't delete or change anything in the excel file"}, status=400)
        
            all_exams = Exam.objects.select_related('student__user').filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, student__in=exams_students, academic_year=current_academic_year, academic_term=term)
            all_exams_data = [{
                'name': f"{x.student.user.first_name} {x.student.user.last_name}",
                'st_id': x.student.st_id,
                'score': x.score,
            } for x in all_exams]
            
            return Response({'message': "Operation successful", 'data': all_exams_data})

        elif request.data['type'] == 'uploadWithoutFile':
            students_class = Classe.objects.get(school=school, name=st_class_name)
            subject_obj = Subject.objects.get(name=subject_name)
            score = float(request.data['score'])
            exams_to_create = []
            for _st_id in json.loads(request.data['selectedStudents']):
                student = Student.objects.get(school=school, st_id=_st_id)
                st_exam = Exam(
                    student=student,
                    subject=subject_obj,
                    teacher=teacher,
                    student_class=students_class,
                    score=score,
                    percentage=0,
                    academic_year=current_academic_year,
                    academic_term=term,
                    school=school,
                )
                exams_to_create.append(st_exam)
            with transaction.atomic():
                try:
                    Exam.objects.bulk_create(exams_to_create)
                    exam_create_obj = Exam.objects.get(school=school, teacher=teacher, student=None, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term)
                    exam_create_obj.delete()
                except Exam.DoesNotExist:
                    transaction.set_rollback(False)
                except Exception:
                    transaction.set_rollback(True)
                    return Response(status=400)
            
            return Response({'message': "Operation successful"})

        elif request.data['type'] == 'editTotalScore':
            new_total_score = Decimal(request.data['totalScore'])
            students_class = Classe.objects.get(school=school, name=st_class_name)
            subject_obj = Subject.objects.get(name=subject_name)
            exams_to_update = []
            results_to_update = []
            grading_system = GradingSystem.objects.filter(schools=school).order_by('lower_limit')
            exams = Exam.objects.filter(school=school, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term)
            for _exam in exams:
                _exam.total_score = new_total_score
                exams_to_update.append(_exam)
                try:
                    st_result = StudentResult.objects.get(school=school, student=_exam.student, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term)
                    new_exam_score = (_exam.score/new_total_score)*st_result.exam_percentage
                    new_exam_score = math.ceil(new_exam_score) if new_exam_score - math.floor(new_exam_score) >= 0.5 else math.floor(new_exam_score)
                    new_result = new_exam_score + st_result.total_assessment_score
                    student_grade = next((item for item in grading_system if item.upper_limit >= new_result >= item.lower_limit), None)
                    st_result.exam_score = new_exam_score
                    st_result.result = new_result
                    st_result.grade = student_grade.label
                    st_result.remark = student_grade.remark
                    results_to_update.append(st_result)
                except StudentResult.DoesNotExist:
                    continue
            with transaction.atomic():
                try:
                    Exam.objects.bulk_update(exams_to_update, ['total_score'])
                    if results_to_update:
                        StudentResult.objects.bulk_update(results_to_update, ['exam_score', 'result', 'grade', 'remark'])
                        all_results =  StudentResult.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).order_by('-result')
                        all_results_data = TeacherGetStudentResultSerializer(all_results, many=True).data
                        return Response(all_results_data)
                except Exception:
                    transaction.set_rollback(True)
                    return Response(status=400)
            return Response()
            
        
        elif request.data['type'] == 'editExamScore':
            new_score = Decimal(request.data['score'])
            with transaction.atomic():
                try:
                    students_class = Classe.objects.get(school=school, name=st_class_name)
                    subject_obj = Subject.objects.get(name=subject_name)
                    student = Student.objects.get(school=school, st_id=request.data['studentId'])
                    st_exam = Exam.objects.get(school=school, student=student, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term)
                    st_exam.score = new_score
                    st_exam.save()
                    st_result = StudentResult.objects.get(school=school, student=student, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term)
                    new_exam_score = (new_score/st_exam.total_score)*st_result.exam_percentage
                    new_exam_score = math.ceil(new_exam_score) if new_exam_score - math.floor(new_exam_score) >= 0.5 else math.floor(new_exam_score)
                    grading_system = GradingSystem.objects.filter(schools=school).order_by('lower_limit')
                    new_result = new_exam_score + st_result.total_assessment_score
                    student_grade = next((item for item in grading_system if item.upper_limit >= new_result >= item.lower_limit), None)
                    st_result.exam_score = new_exam_score
                    st_result.result = new_result
                    st_result.grade = student_grade.label
                    st_result.remark = student_grade.remark
                    st_result.save()
                    return Response({
                        'new_exam_score': new_exam_score, 
                        'new_result': new_result,
                        'new_grade': student_grade.label,
                        'new_remark': student_grade.remark,
                        })
                except StudentResult.DoesNotExist:
                    transaction.set_rollback(False)
                except Exception:
                    transaction.set_rollback(True)
                    return Response(status=400)
            
            return Response()
        
        elif request.data['type'] == 'deleteExam':
            with transaction.atomic():
                try:
                    students_class = Classe.objects.get(school=school, name=st_class_name)
                    subject_obj = Subject.objects.get(name=subject_name)
                    existing_results = StudentResult.objects.filter(school=school, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term).exists()
                    if existing_results:
                        return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
                    exams = Exam.objects.filter(school=school, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term)
                    exams.delete()
                    
                except Exception:
                    transaction.set_rollback(True)
                    return Response(status=400)
            
            return Response()


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def teacher_students_results(request):
    teacher = request.user.staff
    school = teacher.school
    
    if request.method == 'GET':
        current_academic_year = AcademicYear.objects.get(school=school, name=request.GET.get('year'))
        current_term = request.GET.get('term')
        all_results_data = []
        subject_assignments = SubjectAssignment.objects.select_related('teacher', 'students_class').prefetch_related('subjects').filter(school=school, teacher=teacher, academic_year=current_academic_year, academic_term=current_term)
        for assign in subject_assignments:
            results_data = {'class_name': assign.students_class.name, 'results': []}
            for _subject in assign.subjects.all():
                subject_data = {'subject': _subject.name, 'student_results': []}
                all_results = StudentResult.objects.select_related('student__user').filter(school=school, teacher=teacher, subject=_subject, student_class=assign.students_class, academic_year=current_academic_year, academic_term=current_term).order_by('-result')
                if all_results.exists():
                    first_result_data = all_results.first()
                    subject_data['total_assessment_percentage'] = first_result_data.total_assessment_percentage
                    subject_data['exam_percentage'] = first_result_data.exam_percentage
                    subject_data['student_results'] = TeacherGetStudentResultSerializer(all_results, many=True).data
                
                results_data['results'].append(subject_data)
            all_results_data.append(results_data)
            
        return Response(all_results_data)

    elif request.method == 'POST':
        current_academic_year = AcademicYear.objects.get(school=school, name=request.data['year'])
        current_term = int(request.data['term'])
        subject_obj = Subject.objects.get(name=request.data['subject'])
        students_class = Classe.objects.get(school=school, name=request.data['studentsClassName'])
        if request.data['type'] == 'createResults':
            existing_results = StudentResult.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=current_term)
            if existing_results.exists():
                return Response({'message': f"{subject_obj.name} results for this class already exists"}, status=400)
            result_data = json.loads(request.data['resultsData'])
            grading_system = GradingSystem.objects.filter(schools=school).order_by('lower_limit')
            assessments_to_update = []
            exams_to_update = []
            results_to_create = []
            for _student in students_class.students.all():
                student_total_assessment_score = 0
                for _title in result_data['assessments']:
                    student_assessment = Assessment.objects.get(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, student=_student, title=_title['title'], academic_year=current_academic_year, academic_term=current_term)
                    student_total_assessment_score += ((student_assessment.score/_title['totalScore'])*_title['percentage'])
                    student_assessment.percentage = _title['percentage']
                    assessments_to_update.append(student_assessment)
                student_total_assessment_score = math.ceil(student_total_assessment_score) if student_total_assessment_score - math.floor(student_total_assessment_score) >= 0.5 else math.floor(student_total_assessment_score)
                student_exam = Exam.objects.get(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, student=_student, academic_year=current_academic_year, academic_term=current_term)
                student_exam_score = (student_exam.score/student_exam.total_score)*result_data['examsPercentage']
                student_exam_score = math.ceil(student_exam_score) if student_exam_score - math.floor(student_exam_score) >= 0.5 else math.floor(student_exam_score)
                result = student_total_assessment_score + student_exam_score
                student_exam.percentage = result_data['examsPercentage']
                exams_to_update.append(student_exam)
                student_grade = next((item for item in grading_system if item.upper_limit >= result >= item.lower_limit), None)
                student_result = StudentResult.objects.create(
                    school=school,
                    teacher=teacher,
                    student=_student,
                    subject=subject_obj,
                    student_class=students_class,
                    exam_percentage=result_data['examsPercentage'],
                    exam_score=student_exam_score,
                    total_assessment_percentage=result_data['totalAssessmentPercentage'],
                    total_assessment_score=student_total_assessment_score,
                    result=result,
                    remark=student_grade.remark,
                    grade=student_grade.label,
                    academic_year = current_academic_year,
                    academic_term=current_term,
                )
                results_to_create.append(student_result)
            with transaction.atomic():
                try:
                    Assessment.objects.bulk_update(assessments_to_update, ['percentage'])
                    Exam.objects.bulk_update(exams_to_update, ['percentage'])
                    StudentResult.objects.bulk_create(results_to_create)
                except IntegrityError:
                    transaction.set_rollback(False)
                except Exception:
                    transaction.set_rollback(True)
                    return Response(status=400)
                
            subject_data = {'subject': subject_obj.name, 'student_results': []}
            all_results = StudentResult.objects.select_related('student__user').filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=current_term).order_by('-result')
            first_result_data = all_results.first()
            subject_data['total_assessment_percentage'] = first_result_data.total_assessment_percentage
            subject_data['exam_percentage'] = first_result_data.exam_percentage
            subject_data['student_results'] = TeacherGetStudentResultSerializer(all_results, many=True).data
            
            return Response(subject_data)
        
        elif request.data['type'] == 'deleteResults':
            with transaction.atomic():
                results = StudentResult.objects.filter(school=school, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=current_term)
                results.delete()
            
            return Response()


# HOD
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def hod_data(request):
    staff = request.user.staff
    school = staff.school
    
    if request.method == 'GET':
        current_academic_year = AcademicYear.objects.get(school=staff.school, name=request.GET.get('year'))
        current_term = int(request.GET.get('term'))
        department = Department.objects.get(school=school, teachers=staff)
        studentClasses = Classe.objects.filter(school=school, subjects__in=department.subjects.all()).distinct()
        student_classes_data = [x.name for x in studentClasses]
        
        subject_assignments = SubjectAssignment.objects.select_related('teacher', 'students_class').prefetch_related('subjects').filter(school=school, assigned_by=staff, academic_year=current_academic_year, academic_term=current_term)
        subject_assignments_data = HodSubjectAssignmentSerializer(subject_assignments, many=True).data
        
        return Response({
            'student_classes': student_classes_data,
            'subject_assignments': subject_assignments_data,
        })
    
    if request.method == 'POST':
        current_academic_year = AcademicYear.objects.get(school=school, name=request.data['year'])
        current_term = int(request.data['term'])
        students_class = Classe.objects.get(school=school, name=request.data['studentsClassName'])
        teacher = Staff.objects.get(school=school, staff_id=request.data['teacher'])
        if request.data['type'] == 'upload':
            subjects = json.loads(request.data['subjects'])
            subjects_obj = Subject.objects.filter(name__in=subjects).distinct()
            for _subject in subjects_obj:
                if _subject not in teacher.subjects.all():
                    return Response({'message': f"The teacher you selected doesn't teach {_subject.name}"}, status=400)
                elif _subject not in students_class.subjects.all():
                    return Response({'message': f"The class you selected doesn't study {_subject.name}"}, status=400)
            
            with transaction.atomic():
                try:
                    assignment_obj = SubjectAssignment.objects.create(
                        school=school,
                        assigned_by=staff,
                        students_class=students_class,
                        teacher=teacher,
                        academic_year=current_academic_year,
                        academic_term=current_term,
                    )
                    assignment_obj.subjects.set(subjects_obj)
                    assignment_obj.save()
                except IntegrityError:
                    transaction.set_rollback(True)
                    return Response({'message': "Subject assignment with these details already exists"}, status=400)
                except Exception:
                    transaction.set_rollback(True)
                    return Response(status=400)
            
            subject_assignment = SubjectAssignment.objects.get(school=school, assigned_by=staff, students_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=current_term)
            subject_assignments_data = HodSubjectAssignmentSerializer(subject_assignment).data
            return Response(subject_assignments_data)
        
        elif request.data['type'] == 'delete':
            with transaction.atomic():  
                try:
                    existing_exams = Exam.objects.filter(school=school, student_class=students_class, academic_year=current_academic_year, academic_term=current_term).first()
                    if existing_exams:
                        return Response({'message': f"{existing_exams.teacher.user.get_full_name()} has already uploaded exams data for students in the selected class [{students_class.name}]. If you still want to delete this subject assignment, please ask him/her to delete the exams data and try again."}, status=400)
                    existing_assessments = Assessment.objects.select_related('teacher__user').filter(school=school, student_class=students_class, academic_year=current_academic_year, academic_term=current_term).first()
                    if existing_assessments:
                        return Response({'message': f"{existing_assessments.teacher.user.get_full_name()} has already uploaded assessment data for students in the selected class [{students_class.name}]. If you still want to delete this subject assignment, please ask him/her to delete them and try again."}, status=400)
                        
                    assignment_obj = SubjectAssignment.objects.get(
                        school=school,
                        assigned_by=staff,
                        students_class=students_class,
                        teacher=teacher,
                        academic_year=current_academic_year,
                        academic_term=current_term,
                    )
                    assignment_obj.delete()
                except Exception:
                    transaction.set_rollback(True)
                    return Response(status=400)
            
            return Response()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_students_performance(request):
    staff = request.user.staff
    department = DepartmentSerializer(Department.objects.get(school=staff.school, teachers=staff)).data
    academic_year = AcademicYear.objects.get(school=staff.school, name=request.GET.get('year'))

    data = []
    if department:
        for subject in department['subjects']:
            first_year = {'term_one': [], 'term_two': [], 'term_three': []}
            second_year = {'term_one': [], 'term_two': [], 'term_three': []}
            third_year = {'term_one': [], 'term_two': [], 'term_three': []}

            subject_obj = Subject.objects.get(schools=staff.school, name=subject['name'])

            # Year One Student Exams
            result_one_one = ExamsSerializer(
                Exam.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=1,
                    student_year=1,
                ),
                many=True).data
            if result_one_one:
                for result in result_one_one:
                    first_year['term_one'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                first_year['term_one'] = sorted(first_year['term_one'], key=lambda x: x['score'], reverse=True)

            result_one_two = ExamsSerializer(
                Exam.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=2,
                    student_year=1,
                ),
                many=True).data
            if result_one_two:
                for result in result_one_two:
                    first_year['term_two'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                first_year['term_two'] = sorted(first_year['term_two'], key=lambda x: x['score'], reverse=True)

            result_one_three = ExamsSerializer(
                Exam.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=3,
                    student_year=1,
                ),
                many=True).data
            if result_one_three:
                for result in result_one_three:
                    first_year['term_three'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                first_year['term_three'] = sorted(first_year['term_three'], key=lambda x: x['score'], reverse=True)

            # Year Two Students Exams
            result_two_one = ExamsSerializer(
                Exam.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=1,
                    student_year=2,
                ),
                many=True).data
            if result_two_one:
                for result in result_two_one:
                    second_year['term_one'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                second_year['term_one'] = sorted(second_year['term_one'], key=lambda x: x['score'], reverse=True)

            result_two_two = ExamsSerializer(
                Exam.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=2,
                    student_year=2,
                ),
                many=True).data
            if result_two_two:
                for result in result_two_two:
                    second_year['term_two'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                second_year['term_two'] = sorted(second_year['term_two'], key=lambda x: x['score'], reverse=True)

            result_two_three = ExamsSerializer(
                Exam.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=3,
                    student_year=2,
                ),
                many=True).data
            if result_two_three:
                for result in result_two_three:
                    second_year['term_three'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                second_year['term_three'] = sorted(second_year['term_three'], key=lambda x: x['score'], reverse=True)

            # Year Three Students Exams
            result_three_one = ExamsSerializer(
                Exam.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=1,
                    student_year=3,
                ),
                many=True).data
            if result_three_one:
                for result in result_three_one:
                    third_year['term_one'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                third_year['term_one'] = sorted(third_year['term_one'], key=lambda x: x['score'],
                                                reverse=True)

            result_three_two = ExamsSerializer(
                Exam.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=2,
                    student_year=3,
                ),
                many=True).data
            if result_three_two:
                for result in result_three_two:
                    third_year['term_two'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                third_year['term_two'] = sorted(third_year['term_two'], key=lambda x: x['score'],
                                                reverse=True)

            result_three_three = ExamsSerializer(
                Exam.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=3,
                    student_year=3,
                ),
                many=True).data
            if result_three_three:
                for result in result_three_three:
                    third_year['term_three'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                third_year['term_three'] = sorted(third_year['term_three'], key=lambda x: x['score'],
                                                  reverse=True)

            data.append({
                'subject': subject['name'],
                'year_one': first_year,
                'year_two': second_year,
                'year_three': third_year,
            })

        return Response(data)

    else:
        return Response(status=401)