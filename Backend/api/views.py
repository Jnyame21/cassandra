# Django
import json
import os
import random

from django.contrib.auth.models import User
from django.http import HttpResponse
from django.http import HttpResponse
from django.db import IntegrityError, transaction
from django.db.utils import IntegrityError as IntergrityError_unique_constraint
from django.core.files.storage import default_storage
from django.core.exceptions import SuspiciousOperation
from django.core.files.base import ContentFile
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.core.mail import send_mail
from django.shortcuts import get_object_or_404

# Django Restframework
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.generics import ListAPIView
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework.permissions import IsAuthenticated

# Other
# from api.models import *
# from api.serializer import *
from datetime import datetime, date
import io
from pprint import pprint
import pytz
from dateutil import parser
from api.utils import *

# Document Manipulation
import pandas as pd
from openpyxl.styles import Font, Alignment, Protection
from openpyxl import Workbook, load_workbook


# LOGIN
class UserAuthSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)

        # Add custom claims
        user_data = UserSerializer(user).data
        try:
            staff = user.staff
            staff_data = SpecificStaffSerializer(staff).data
            if staff_data['contact'] and staff_data['contact'] != 'not set':
                token['reset'] = False
            else:
                token['reset'] = True

            return token

        except User.staff.RelatedObjectDoesNotExist:
            try:
                student = user.student
                if user_data['last_login']:
                    token['last_login'] = 'yes'

                return token

            except User.student.RelatedObjectDoesNotExist:
                try:
                    head = user.head
                    head_data = SpecificHeadSerializer(head).data
                    if head_data['contact'] != 'not set':
                        token['reset'] = False
                    else:
                        token['reset'] = True

                    return token

                except User.head.RelatedObjectDoesNotExist:
                    return token


class UserAuthView(TokenObtainPairView):
    serializer_class = UserAuthSerializer


# User Data
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def get_user_data(request):
    user = request.user
    user_info = UserSerializer(user).data

    if request.method == 'POST':
        data = request.data

        if len(data['contact']) != 10:
            return Response({'ms': 'The contact must be a 10 digit number'}, status=201)

        elif data['contact'][0] != '0' or data['altContact'] and data['altContact'] != 'null' and data['altContact'][0] != '0':
            return Response({'ms': 'The phone numbers must start with a 0'}, status=201)

        elif data['altContact'] and data['altContact'] != 'null' and len(data['altContact']) != 10:
            return Response({'ms': 'The second phone number must be a 10 digit number'}, status=201)

        try:
            staff = user.staff
            with transaction.atomic():
                user.set_password(data['password'])
                user.email = data['email'] if data['email'] and data['email'] != 'null' else 'not set'
                user.save()
                staff.contact = data['contact']
                staff.alt_contact = data['altContact'] if data['altContact'] and data['altContact'] != 'null' else 'not set'
                staff.save()

            return Response(status=200)

        except User.staff.RelatedObjectDoesNotExist:
            try:
                head = user.head
                with transaction.atomic():
                    user.set_password(data['password'])
                    user.email = data['email'].strip() if data['email'] and data['email'] != 'null' else 'not set'
                    user.save()
                    head.contact = data['contact']
                    head.alt_contact = data['altContact'] if data['altContact'] and data['altContact'] != 'null' else 'not set'
                    head.save()

                return Response(status=200)

            except User.head.RelatedObjectDoesNotExist:
                return Response(status=401)

    else:
        user_data = {
            'username': user_info['username'],
            'first_name': user_info['first_name'],
            'last_name': user_info['last_name']
        }
        if user_info['email']:
            user_data['email'] = user_info['email']
        else:
            user_data['email'] = 'null'

        try:
            staff = user.staff
            # Get current academic year
            get_current_academic_year(staff, user_data)

            staff_data = StaffSerializer(staff).data
            user_data['last_login'] = format_relative_date_time(staff_data['user']['last_login'], False, True)
            user_data['role'] = 'staff'
            user_data['img'] = staff_data['img']
            user_data['school'] = staff_data['school']
            user_data['staff_id'] = staff_data['staff_id']
            try:
                department = Department.objects.get(teachers=staff)
                user_data['department'] = DepartmentNameSerializer(department).data['name']
                if department.hod == staff:
                    user_data['staff_role'] = 'hod'
                    staff.role = 'hod'
                    staff.save()
                else:
                    user_data['staff_role'] = 'teacher'

            except Department.DoesNotExist:
                user_data['staff_role'] = staff_data['role']

            user_data['subjects'] = staff_data['subjects']
            user_data['gender'] = staff_data['gender']
            user_data['dob'] = staff_data['dob']
            user_data['address'] = staff_data['address']
            user_data['contact'] = staff_data['contact']
            user_data['ms'] = 'Login successful'
            user_data['pob'] = staff_data['pob']
            user_data['region'] = staff_data['region']
            user_data['nationality'] = staff_data['nationality']

            return Response(user_data)

        except User.staff.RelatedObjectDoesNotExist:
            try:
                student = user.student

                # Get current academic year
                get_current_academic_year(student, user_data)

                student_data = StudentSerializer(student).data
                user_data['last_login'] = format_relative_date_time(student_data['user']['last_login'], False, True)
                user_data['role'] = 'student'

                user_data['img'] = student_data['img']
                user_data['ms'] = 'Login successful'
                user_data['program'] = student_data['program']['name']
                user_data['school'] = student_data['school']
                user_data['st_id'] = student_data['st_id']

                if student_data['index_no']:
                    user_data['index_no'] = student_data['index_no']
                else:
                    user_data['index_no'] = 'not assigned yet'

                try:
                    clas = Classe.objects.get(school=student.school, students=student)
                    class_name = ClasseWithoutStudentsSerializer(clas).data
                    user_data['st_class'] = class_name['name']

                    if student_data['has_completed']:
                        user_data['current_yr'] = 'COMPLETED'
                    else:
                        current_date = timezone.now().date()
                        if current_date >= clas.completion_date:
                            student.has_completed = True
                            student.current_year = clas.students_year
                            student.save()
                            user_data['current_yr'] = 'COMPLETED'

                        else:
                            user_data['current_yr'] = student_data['current_year']

                except Classe.DoesNotExist:
                    pass

                user_data['contact'] = student_data['contact']
                user_data['dob'] = student_data['dob']
                user_data['gender'] = student_data['gender']
                user_data['pob'] = student_data['pob']
                user_data['region'] = student_data['region']
                user_data['address'] = student_data['address']
                user_data['religion'] = student_data['religion']
                user_data['nationality'] = student_data['nationality']
                user_data['guardian'] = student_data['guardian']
                user_data['guardian_gender'] = student_data['guardian_gender']
                user_data['guardian_occupation'] = student_data['guardian_occupation']
                user_data['guardian_nationality'] = student_data['guardian_nationality']

                if student_data['guardian_email']:
                    user_data['guardian_email'] = student_data['guardian_email']
                else:
                    user_data['guardian_email'] = 'null'

                user_data['guardian_contact'] = student_data['guardian_contact']
                user_data['guardian_address'] = student_data['guardian_address']

                return Response(user_data)

            except User.student.RelatedObjectDoesNotExist:
                head = user.head
                # Get current academic year
                get_current_academic_year(head, user_data)

                head_data = HeadSerializer(head).data
                user_data['last_login'] = format_relative_date_time(head_data['user']['last_login'], False, True)
                user_data['role'] = 'head'

                user_data['img'] = head_data['img']
                user_data['school'] = head_data['school']
                user_data['head_id'] = head_data['head_id']
                user_data['head_role'] = head_data['role']
                user_data['gender'] = head_data['gender']
                user_data['dob'] = head_data['dob']
                user_data['address'] = head_data['address']
                user_data['contact'] = head_data['contact']
                user_data['ms'] = 'Login successful'
                user_data['pob'] = head_data['pob']
                user_data['region'] = head_data['region']
                user_data['nationality'] = head_data['nationality']

                return Response(user_data)


# STUDENT
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def get_student_data(request):
    school = request.user.student.school
    if request.method == 'GET':
        academic_year = AcademicYear.objects.get(school=school, name=request.GET.get('year'))
        term = request.GET.get('term')
        student = request.user.student
        if AcademicYearSerializer(academic_year).data not in AcademicYearSerializer(student.st_class.academic_years.all(), many=True).data:
            years = AcademicYearSerializer(student.st_class.academic_years, many=True).data
            if len(years) == 1:
                academic_year = AcademicYear.objects.get(school=school, name=years[0]['name'])
            elif len(years) > 1:
                years = sorted(years, key=lambda x: x['start_date'])
                academic_year = AcademicYear.objects.get(school=school, name=years[-1]['name'])

        student_class = Classe.objects.get(students=student)

        results_term_one = StudentResultsSerializer(
            Result.objects.filter(academic_year=academic_year, student=student, academic_term=1), many=True).data
        results_term_two = StudentResultsSerializer(
            Result.objects.filter(academic_year=academic_year, student=student, academic_term=2), many=True).data
        results_term_three = StudentResultsSerializer(
            Result.objects.filter(academic_year=academic_year, student=student, academic_term=3), many=True).data

        student_results = {'term_one': results_term_one, 'term_two': results_term_two, 'term_three': results_term_three}

        class_data = ClasseWithSubjectsSerializer(student_class).data
        student_subjects = []
        for subject in class_data['subjects']:
            subject_data = {'name': subject['name']}
            st_subject = Subject.objects.get(name=subject['name'])
            subject_assignment = SubjectAssignment.objects.filter(academic_year=academic_year, academic_term=term,
                                                                students_class=student_class, subject=st_subject).first()
            if subject_assignment:
                subject_assignment_data = SubjectAssignmentSerializer(subject_assignment).data
                subject_data['teacher'] = f"{ subject_assignment_data['teacher']['user']['first_name']} {subject_assignment_data['teacher']['user']['last_name']}"
                subject_data['teacher_img'] = f"{subject_assignment_data['teacher']['img']}"
                subject_data['teacher_contact'] = f"{subject_assignment_data['teacher']['contact']}"
                subject_data['teacher_email'] = f"{subject_assignment_data['teacher']['user']['email']}"
                subject_data['teacher_gender'] = f"{subject_assignment_data['teacher']['gender']}"
                subject_data['teacher_department'] = f"{subject_assignment_data['teacher']['department']['name']}"

            else:
                subject_data['teacher'] = 'null'
                subject_data['teacher_img'] = 'null'
                subject_data['teacher_contact'] = 'null'
                subject_data['teacher_email'] = 'null'
                subject_data['teacher_gender'] = 'null'
                subject_data['teacher_department'] = 'null'

            student_subjects.append(subject_data)
        if len(class_data['academic_years']) > 1:
            return Response({'subjects': student_subjects, 'results': student_results,
                             'academic_years': sorted(class_data['academic_years'], key=lambda x: x['start_date'], reverse=True),
                             'class_data': class_data['students']
                             })
        else:
            return Response({'subjects': student_subjects, 'results': student_results, 'class_data': class_data['students']})

    else:
        academic_year = AcademicYear.objects.get(school=school, name=request.data['year'])
        student = request.user.student

        results_term_one = StudentResultsSerializer(
            Result.objects.filter(academic_year=academic_year, student=student, academic_term=1), many=True).data
        results_term_two = StudentResultsSerializer(
            Result.objects.filter(academic_year=academic_year, student=student, academic_term=2), many=True).data
        results_term_three = StudentResultsSerializer(
            Result.objects.filter(academic_year=academic_year, student=student, academic_term=3), many=True).data

        return Response({'term_one': results_term_one, 'term_two': results_term_two, 'term_three': results_term_three})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_transcript(request):
    file_path = get_student_transcript(request.user.student)
    return Response(file_path)


# TEACHER
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def teacher_students_attendance(request):
    teacher = request.user.staff
    if request.method == 'POST':
        data = request.data
        academic_year = AcademicYear.objects.get(school=teacher.school, name=data['year'])
        subject = Subject.objects.get(schools=teacher.school, name=data['subject'])
        students_class = Classe.objects.prefetch_related('students').get(school=teacher.school, name=data['className'])
        if data['type'] == 'create':
            students = data['absentStudents'].split(',')
            try:
                existing_attendance = StudentAttendance.objects.get(
                    school=teacher.school,
                    date=data['date'],
                    teacher=teacher,
                    subject=subject,
                    students_class=students_class,
                    academic_year=academic_year,
                    academic_term=int(data['term']),
                )
                return Response({
                    'ms': f"You have already uploaded attendance for the date {data['date']}",
                }, status=201)

            except StudentAttendance.DoesNotExist:
                with transaction.atomic():
                    attendance = StudentAttendance.objects.create(
                        school=teacher.school,
                        subject=subject,
                        teacher=teacher,
                        date=data['date'],
                        students_class=students_class,
                        academic_year=academic_year,
                        academic_term=int(data['term']),
                        students_year=students_class.students_year,
                    )
                    for st in students_class.students.all():
                        if st.st_id not in students:
                            attendance.students_present.add(st)
                        else:
                            attendance.students_absent.add(st)

                    attendance.save()

                new_attendance = StudentsAttendanceSerializer(StudentAttendance.objects.get(
                    school=teacher.school,
                    academic_year=academic_year,
                    academic_term=int(data['term']),
                    students_year=students_class.students_year,
                    subject=subject,
                    date=data['date'],
                    students_class=students_class,
                    teacher=teacher
                )).data

                return Response(new_attendance)

        elif data['type'] == 'delete':
            try:
                students_class = Classe.objects.get(school=teacher.school, name=data['className'])
                with transaction.atomic():
                    student_attendance = StudentAttendance.objects.get(
                        school=teacher.school,
                        subject=subject,
                        academic_year=academic_year,
                        academic_term=int(data['term']),
                        students_class=students_class,
                        teacher=teacher,
                        date=data['date'],
                    )
                    student_attendance.delete()

                return Response(status=200)

            except StudentAttendance.DoesNotExist:
                return Response(status=401)

    else:
        academic_year = AcademicYear.objects.get(school=teacher.school, name=request.GET.get('year'))
        term = request.GET.get('term')
        subject_assignments = SubjectAssignment.objects.prefetch_related('subject').filter(
            school=teacher.school,
            academic_year=academic_year,
            academic_term=term,
            teacher=teacher,
        )
        students_attendance = []
        if subject_assignments.exists():
            for assign in subject_assignments:
                assign_data = SubjectAssignmentWithoutStudentsSerializer(assign).data
                attendance = StudentsAttendanceSerializer(StudentAttendance.objects.filter(
                    school=teacher.school,
                    academic_year=academic_year,
                    academic_term=term,
                    students_class=assign.students_class,
                    subject=assign.subject
                ).order_by('-date'), many=True).data

                students_attendance.append({
                    'subject': assign_data['subject']['name'],
                    'class_name': assign_data['students_class']['name'],
                    'students_year': assign_data['students_class']['students_year'],
                    'attendance': attendance,
                })

        return Response(students_attendance)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_teacher_subject_assignments(request):
    sch = request.user.staff.school
    try:
        year = request.GET.get('year')
        teacher = request.user.staff
        academic_year = AcademicYear.objects.get(school=sch, name=year)
        term1_subjects_assignments = SubjectAssignmentSerializer(
            SubjectAssignment.objects.filter(teacher=teacher, academic_year=academic_year, academic_term=1), many=True)
        term2_subjects_assignments = SubjectAssignmentSerializer(
            SubjectAssignment.objects.filter(teacher=teacher, academic_year=academic_year, academic_term=2), many=True)
        term3_subjects_assignments = SubjectAssignmentSerializer(
            SubjectAssignment.objects.filter(teacher=teacher, academic_year=academic_year, academic_term=3), many=True)
        return Response(
            {'subject_assignments': {
                'term1': term1_subjects_assignments.data, 
                'term2': term2_subjects_assignments.data, 
                'term3': term3_subjects_assignments.data
                }
            }
        )
    except request.user.staff.DoesNotExist:
        return Response(status=401)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def teacher_student_results(request):
    sch = request.user.staff.school
    try:
        teacher = request.user.staff
        staff = StaffSerializer(teacher).data
        if request.method == 'POST':
            year = request.data['year']
            term = int(request.data['term'])
            subject = request.data['subject']
            academic_year = AcademicYear.objects.get(school=sch, name=year)
            if request.data['score'] == 'generate':
                st_class = request.data['students_class']
                st_year = request.data['students_year']
                students = request.data['st_id']
                data = [
                    ['STUDENT NAME', 'STUDENT ID', 'SCORE']
                ]

                for st in students:
                    row = [st['name'], st['st_id'], 'not set']
                    data.append(row)

                wb = Workbook()
                ws = wb.active
                ws.title = subject

                ws.merge_cells('A1:I3')

                if sch.semesters:
                    ws['A1'].value = f"SUBJECT: [{subject}] - ACADEMIC YEAR: [{year}] - CLASS: [{st_class} FORM {st_year}] - SEMESTER: [{term}]"
                else:
                    ws['A1'].value = f"SUBJECT: [{subject}] - ACADEMIC YEAR: [{year}] - CLASS: [{st_class} FORM {st_year}] - TRIMESTER: [{term}]"

                ws['A1'].font = Font(size=14, bold=True)
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

                count = 0
                container = ws['B5:D200']
                while count < len(data):
                    for row_ws, row_data in zip(container, data):
                        for cell, item in zip(row_ws, row_data):
                            cell.value = item

                    count += 1

                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 25

                for row in ws['A3:F200']:
                    for cell in row:
                        cell.font = Font(size=14)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                for row in ws['B3:B200']:
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

                for cells in ws['A5:F200'][0]:
                    cells.font = Font(bold=True, size=14)

                ws.protection.sheet = True
                for cell in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=4, max_col=4):
                    cell[0].protection = Protection(locked=False)

                ws.protection.password = 'teamjn'

                filename = f"{subject.replace(' ', '-')}-{st_class}-FORM-{st_year}-Results.xlsx"
                byte_file = io.BytesIO()
                wb.save(byte_file)

                if settings.DEBUG:
                    file_path = f"{get_school_folder(staff['school']['name'])}/staff/{staff['user']['username']}/{filename}"
                    if default_storage.exists(file_path):
                        default_storage.delete(file_path)

                    save_file = default_storage.save(file_path, byte_file)

                    return Response({
                        'filename': filename,
                        'message': 'File generated successfully',
                        'file_path': f"http://localhost:8000{default_storage.url(save_file)}",
                    })

                else:
                    file_path = f"media/{get_school_folder(staff['school']['name'])}/staff/{staff['user']['username']}/{filename}"
                    if default_storage.exists(file_path):
                        default_storage.delete(file_path)

                    save_file = default_storage.save(file_path, byte_file)

                    return Response({
                        'filename': filename,
                        'message': 'File generated successfully',
                        'file_path': default_storage.url(save_file),
                    })

            elif request.data['score'] == 'upload':
                wb = load_workbook(request.data['st_id'])
                ws = wb.active
                ws.protection = False

                if sch.semesters:
                    if ws[
                        'A1'].value != f"SUBJECT: [{subject}] - ACADEMIC YEAR: [{year}] - CLASS: [{request.data['students_class']} FORM {request.data['students_year']}] - SEMESTER: [{term}]":
                        return Response({'message': "Invalid file. You must upload the same file you generated."},
                                        status=201)
                else:
                    if ws[
                        'A1'].value != f"SUBJECT: [{subject}] - ACADEMIC YEAR: [{year}] - CLASS: [{request.data['students_class']} FORM {request.data['students_year']}] - TRIMESTER: [{term}]":
                        return Response({'message': "Invalid file. You must upload the same file you generated."},
                                        status=201)

                df = pd.read_excel(request.data['st_id'])
                uncleaned_data = df.values.tolist()
                first_clean_data = []
                for row in uncleaned_data:
                    first_clean_row = [item for item in row if not pd.isna(item)]
                    if first_clean_row:
                        first_clean_data.append(first_clean_row)

                first_clean_data.pop(0)
                if not first_clean_data:
                    return Response({'message': f"There are no students data in the uploaded file"}, status=201)

                students_results = [[float(item) if isinstance(item, (int, float)) else item for item in row] for row in
                                    first_clean_data]
                for row in students_results:
                    for item in row:
                        if row.index(item) == 2 and not isinstance(item, float):
                            return Response({
                                'message': f"You have not assigned a score to {row[0]}"},
                                status=201)
                        elif row.index(item) == 2 and item > 100:
                            return Response({
                                'message': f"Scores cannot be more than 100. Recheck the score for {row[0]} in the file"},
                                status=201)
                        elif row.index(item) == 2 and item < 0:
                            return Response({
                                'message': f"Scores cannot be negative. Recheck the score for {row[0]} in the file"},
                                status=201)

                students_id = []
                students_score = []
                for row in students_results:
                    for item in row:
                        if row.index(item) == 1:
                            students_id.append(item)
                        if row.index(item) == 2:
                            students_score.append(item)

                students_class = Classe.objects.get(
                    school=sch,
                    name=request.data['students_class'],
                    students_year=request.data['students_year']
                )
                with transaction.atomic():
                    for st_id, st_score in zip(students_id, students_score):
                        if students_class.students.filter(st_id=st_id).exists():
                            st_subject = Subject.objects.get(name=subject)
                            if students_class.subjects.filter(name=st_subject.name).exists():
                                st = Student.objects.get(st_id=st_id)
                                try:
                                    st_result = Result.objects.create(
                                        student=st,
                                        subject=st_subject,
                                        teacher=teacher,
                                        score=st_score,
                                        student_year=st.current_year,
                                        academic_year=academic_year,
                                        academic_term=term,
                                        school=sch
                                    )
                                    st_result.save()

                                except IntegrityError:
                                    return Response(
                                        {
                                            'message': f"You have already uploaded result for student with ID {st.st_id}"
                                        }, status=201)

                            else:
                                transaction.set_rollback(True)
                                return Response(
                                    {
                                        'message': f"Student does not study {subject}"
                                    },
                                    status=201)

                        else:
                            transaction.set_rollback(True)
                            return Response(
                                {
                                    'message': f"Students must be in {request.data['students_class']} {request.data['students_year']}"
                                },
                                status=201)

                    data = teacher_results_upload(teacher, academic_year, term)
                    return Response({'data': data, 'message': "Results uploaded and saved successfully"}, status=200)

            elif request.data['score'] == 'edit':
                student_classe = Classe.objects.get(
                    school=sch, name=request.data['student_class'],
                    students_year=request.data['student_year']
                )
                student_subject = Subject.objects.get(schools=sch, name=request.data['subject'])
                st = student_classe.students.get(st_id=request.data['st_id'])
                result = Result.objects.get(
                    student=st,
                    teacher=teacher,
                    subject=student_subject,
                    academic_year=academic_year,
                    academic_term=term,
                    school=sch
                )
                with transaction.atomic():
                    try:
                        mark = float(request.data['mark'])
                        result.score = mark
                        result.save()

                    except ValueError:
                        return Response({'message': "The score must be a number"}, status=202)

                data = teacher_results_upload(teacher, academic_year, term)
                return Response(data, status=200)

            else:
                score = float(request.data['score'])
                academic_year = AcademicYear.objects.get(school=sch, name=year)

                if isinstance(request.data['st_id'], list):
                    students_id = request.data['st_id']
                    for st_id in students_id:
                        student = Student.objects.get(school=sch, st_id=st_id)
                        if student.program.subjects.filter(name=subject).exists():
                            subject_obj = Subject.objects.get(name=subject)
                            new_result = Result.objects.create(
                                student=student,
                                subject=subject_obj,
                                student_year=student.current_year,
                                teacher=teacher,
                                score=score,
                                academic_year=academic_year,
                                academic_term=term,
                                school=sch
                            )
                            new_result.save()

                        else:
                            return Response(status=404)

                    response_data = teacher_results_upload(teacher, academic_year, term)
                    return Response(response_data)

                elif isinstance(request.data['st_id'], str):
                    student_id = request.data['st_id']
                    student = Student.objects.get(st_id=student_id)
                    if student.program.subjects.filter(name=subject).exists():
                        subject_obj = Subject.objects.get(name=subject)
                        new_result = Result.objects.create(
                            student=student,
                            subject=subject_obj,
                            teacher=teacher,
                            score=score,
                            student_year=student.current_year,
                            academic_year=academic_year,
                            academic_term=term,
                            school=sch
                        )
                        new_result.save()

                        response_data = teacher_results_upload(teacher, academic_year, term)
                        return Response(response_data)

                    else:
                        return Response(status=404)
                else:
                    return Response(status=401)

        else:
            year = request.GET.get('year')
            term = request.GET.get('term')
            academic_year = AcademicYear.objects.get(school=sch, name=year)
            response_data = teacher_results_upload(teacher, academic_year, term)

            return Response(response_data)

    except request.user.staff.DoesNotExist:
        return Response(status=401)


# HOD
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def get_hod_data(request):
    staff = request.user.staff
    # Hod crate subject assignments
    if request.method == 'POST' and request.data['type'] == 'create':
        subject = Subject.objects.get(schools=staff.school, name=request.data['subject'])
        class_name = request.data['class_name'].split(' ', 1)
        students_class = Classe.objects.get(school=staff.school, name=class_name[0])
        academic_year = AcademicYear.objects.get(school=staff.school, name=request.data['year'])
        teacher = Staff.objects.get(school=staff.school, staff_id=request.data['staff_id'])
        if teacher.subjects.filter(name=subject.name).exists():
            if students_class.subjects.filter(name=subject.name).exists():
                with transaction.atomic():
                    try:
                        subjects_assignment = SubjectAssignment.objects.create(
                            hod=staff,
                            subject=subject,
                            students_class=students_class,
                            teacher=teacher,
                            academic_year=academic_year,
                            academic_term=int(request.data['term']),
                            school=staff.school
                        )
                        subjects_assignment.save()
                        response_data = get_hod_subject_assignments(staff, academic_year)
                        return Response(
                            {
                                'ms': f"subject assignment uploaded and saved successfully",
                                'data': response_data},
                            status=200)

                    except IntegrityError:
                      return Response({'ms': "Subject assignment with these details already exists"}, status=201)

            else:
                return Response({'ms': f"The class you selected don't study {request.data['subject']}"}, status=201 )
        else:
            return Response({'ms': f"The teacher you selected doesn't teach {request.data['subject']}"}, status=201)

        # Hod delete subject assignment
    elif request.method == 'POST' and request.data['type'] == 'delete':
        subject = Subject.objects.get(schools=staff.school, name=request.data['subject_name'])
        students_class = Classe.objects.get(school=staff.school, name=request.data['students_class_name'])
        academic_year = AcademicYear.objects.get(school=staff.school, name=request.data['year'])
        teacher = Staff.objects.get(school=staff.school, staff_id=request.data['teacher_id'])

        subjects_assignment = SubjectAssignment.objects.get(
            subject=subject,
            hod=staff,
            students_class=students_class,
            teacher=teacher,
            academic_year=academic_year,
            academic_term=int(request.data['term']),
            school=staff.school
        )
        subjects_assignment.delete()

        return Response(status=200)

    else:
        academic_year = AcademicYear.objects.get(school=staff.school, name=request.GET.get('year'))
        response_data = get_hod_subject_assignments(staff, academic_year)
        return Response(response_data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_students_performance(request):
    staff = request.user.staff
    department = DepartmentSerializer(Department.objects.get(school=staff.school, teachers=staff)).data
    academic_year = AcademicYear.objects.get(school=staff.school, name=request.GET.get('year'))

    data = []
    if department:
        for subject in department['subjects']:
            first_year = {'term_one': [], 'term_two': [], 'term_three': []}
            second_year = {'term_one': [], 'term_two': [], 'term_three': []}
            third_year = {'term_one': [], 'term_two': [], 'term_three': []}

            subject_obj = Subject.objects.get(schools=staff.school, name=subject['name'])

            # Year One Student Results
            result_one_one = ResultsSerializer(
                Result.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=1,
                    student_year=1,
                ),
                many=True).data
            if result_one_one:
                for result in result_one_one:
                    first_year['term_one'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                first_year['term_one'] = sorted(first_year['term_one'], key=lambda x: x['score'], reverse=True)

            result_one_two = ResultsSerializer(
                Result.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=2,
                    student_year=1,
                ),
                many=True).data
            if result_one_two:
                for result in result_one_two:
                    first_year['term_two'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                first_year['term_two'] = sorted(first_year['term_two'], key=lambda x: x['score'], reverse=True)

            result_one_three = ResultsSerializer(
                Result.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=3,
                    student_year=1,
                ),
                many=True).data
            if result_one_three:
                for result in result_one_three:
                    first_year['term_three'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                first_year['term_three'] = sorted(first_year['term_three'], key=lambda x: x['score'], reverse=True)

            # Year Two Students Results
            result_two_one = ResultsSerializer(
                Result.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=1,
                    student_year=2,
                ),
                many=True).data
            if result_two_one:
                for result in result_two_one:
                    second_year['term_one'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                second_year['term_one'] = sorted(second_year['term_one'], key=lambda x: x['score'], reverse=True)

            result_two_two = ResultsSerializer(
                Result.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=2,
                    student_year=2,
                ),
                many=True).data
            if result_two_two:
                for result in result_two_two:
                    second_year['term_two'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                second_year['term_two'] = sorted(second_year['term_two'], key=lambda x: x['score'], reverse=True)

            result_two_three = ResultsSerializer(
                Result.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=3,
                    student_year=2,
                ),
                many=True).data
            if result_two_three:
                for result in result_two_three:
                    second_year['term_three'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                second_year['term_three'] = sorted(second_year['term_three'], key=lambda x: x['score'], reverse=True)

            # Year Three Students Results
            result_three_one = ResultsSerializer(
                Result.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=1,
                    student_year=3,
                ),
                many=True).data
            if result_three_one:
                for result in result_three_one:
                    third_year['term_one'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                third_year['term_one'] = sorted(third_year['term_one'], key=lambda x: x['score'],
                                                reverse=True)

            result_three_two = ResultsSerializer(
                Result.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=2,
                    student_year=3,
                ),
                many=True).data
            if result_three_two:
                for result in result_three_two:
                    third_year['term_two'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                third_year['term_two'] = sorted(third_year['term_two'], key=lambda x: x['score'],
                                                reverse=True)

            result_three_three = ResultsSerializer(
                Result.objects.filter(
                    school=staff.school,
                    academic_year=academic_year,
                    subject=subject_obj,
                    academic_term=3,
                    student_year=3,
                ),
                many=True).data
            if result_three_three:
                for result in result_three_three:
                    third_year['term_three'].append({
                        'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                        'st_id': result['student']['st_id'],
                        'score': float(result['score']),
                    })

                third_year['term_three'] = sorted(third_year['term_three'], key=lambda x: x['score'],
                                                  reverse=True)

            data.append({
                'subject': subject['name'],
                'year_one': first_year,
                'year_two': second_year,
                'year_three': third_year,
            })

        return Response(data)

    else:
        return Response(status=401)


# HEAD
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_head_data(request):
    head = request.user.head
    try:
        data = []
        academic_year = AcademicYear.objects.get(school=head.school, name=request.GET.get('year'))
        departments = DepartmentSerializer(Department.objects.filter(school=head.school), many=True).data

        for department in departments:
            subject_assignments = {'term_one': [], 'term_two': [], 'term_three': []}
            for subject in department['subjects']:
                subject_obj = Subject.objects.get(schools=head.school, name=subject['name'])
                for teacher in department['teachers']:
                    teacher_obj = Staff.objects.get(staff_id=teacher['staff_id'])

                    term_one_subject_assignments = SubjectAssignment.objects.filter(
                        academic_year=academic_year,
                        teacher=teacher_obj,
                        subject=subject_obj,
                        academic_term=1,
                        school=head.school
                    )
                    if term_one_subject_assignments.exists():
                        serialized_data = SubjectAssignmentWithoutStudentsSerializer(term_one_subject_assignments, many=True).data
                        if len(serialized_data) == 1:
                            subject_assignments['term_one'].append(serialized_data[0])
                        elif len(serialized_data) > 1:
                            for item in serialized_data:
                                subject_assignments['term_one'].append(item)

                    term_two_subject_assignments = SubjectAssignment.objects.filter(
                        academic_year=academic_year,
                        teacher=teacher_obj,
                        subject=subject_obj,
                        academic_term=2,
                        school=head.school
                    )

                    if term_two_subject_assignments.exists():
                        serialized_data = SubjectAssignmentWithoutStudentsSerializer(term_two_subject_assignments,
                                                                                     many=True).data
                        if len(serialized_data) == 1:
                            subject_assignments['term_two'].append(serialized_data[0])
                        elif len(serialized_data) > 1:
                            subject_assignments['term_two'].extend(serialized_data)

                    term_three_subject_assignments = SubjectAssignment.objects.filter(
                        academic_year=academic_year,
                        teacher=teacher_obj,
                        subject=subject_obj,
                        academic_term=3,
                        school=head.school
                    )

                    if term_three_subject_assignments.exists():
                        serialized_data = SubjectAssignmentWithoutStudentsSerializer(term_three_subject_assignments,
                                                                                     many=True).data
                        if len(serialized_data) == 1:
                            subject_assignments['term_three'].append(serialized_data[0])
                        elif len(serialized_data) > 1:
                            subject_assignments['term_three'].extend(serialized_data)

            data.append({
                'department': department,
                'subject_assignments': subject_assignments,
            })

        programs = []

        sch_programs = Program.objects.filter(schools=head.school)
        for program in sch_programs:
            program_data = {'name': ProgramSerializer(program).data['name'], 'classes': []}
            classes = ClasseSerializer(Classe.objects.filter(school=head.school, program=program, is_active=True), many=True).data
            for clas in classes:
                program_data['classes'].append(clas)

            program_data['classes'] = sorted(program_data['classes'], key=lambda x: x['students_year'])
            programs.append(program_data)

        response_data = {
            'departments': data,
            'programs': programs,
        }

        return Response(response_data)
    except Exception as e:
        pass


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def head_students_performance(request):
    head = request.user.head
    departments = DepartmentSerializer(Department.objects.filter(school=head.school, ), many=True).data
    academic_year = AcademicYear.objects.get(school=head.school, name=request.GET.get('year'))

    data = []
    if departments:
        for department in departments:
            first_year = {'term_one': [], 'term_two': [], 'term_three': []}
            second_year = {'term_one': [], 'term_two': [], 'term_three': []}
            third_year = {'term_one': [], 'term_two': [], 'term_three': []}

            for subject in department['subjects']:
                subject_obj = Subject.objects.get(name=subject['name'])

                # Year One Student Results
                result_one_one = ResultsSerializer(
                    Result.objects.filter(
                        school=head.school,
                        academic_year=academic_year,
                        subject=subject_obj,
                        academic_term=1,
                        student_year=1,
                    ),
                    many=True).data
                if result_one_one:
                    for result in result_one_one:
                        first_year['term_one'].append({
                            'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                            'st_id': result['student']['st_id'],
                            'score': float(result['score']),
                        })

                    first_year['term_one'] = sorted(first_year['term_one'], key=lambda x: float(x['score']), reverse=True)

                result_one_two = ResultsSerializer(
                    Result.objects.filter(
                        school=head.school,
                        academic_year=academic_year,
                        subject=subject_obj,
                        academic_term=2,
                        student_year=1,
                    ),
                    many=True).data
                if result_one_two:
                    for result in result_one_two:
                        first_year['term_two'].append({
                            'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                            'st_id': result['student']['st_id'],
                            'score': float(result['score']),
                        })

                    first_year['term_two'] = sorted(first_year['term_two'], key=lambda x: float(x['score']), reverse=True)

                result_one_three = ResultsSerializer(
                    Result.objects.filter(
                        school=head.school,
                        academic_year=academic_year,
                        subject=subject_obj,
                        academic_term=3,
                        student_year = 1,
                    ),
                    many=True).data
                if result_one_three:
                    for result in result_one_three:
                        first_year['term_three'].append({
                            'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                            'st_id': result['student']['st_id'],
                            'score': float(result['score']),
                        })

                    first_year['term_three'] = sorted(first_year['term_three'], key=lambda x: float(x['score']), reverse=True)

                # Year Two Students Results
                result_two_one = ResultsSerializer(
                    Result.objects.filter(
                        school=head.school,
                        academic_year=academic_year,
                        subject=subject_obj,
                        academic_term=1,
                        student_year=2,
                    ),
                    many=True).data
                if result_two_one:
                    for result in result_two_one:
                        second_year['term_one'].append({
                            'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                            'st_id': result['student']['st_id'],
                            'score': float(result['score']),
                        })

                    second_year['term_one'] = sorted(second_year['term_one'], key=lambda x: float(x['score']), reverse=True)

                result_two_two = ResultsSerializer(
                    Result.objects.filter(
                        school=head.school,
                        academic_year=academic_year,
                        subject=subject_obj,
                        academic_term=2,
                        student_year=2,
                    ),
                    many=True).data
                if result_two_two:
                    for result in result_two_two:
                        second_year['term_two'].append({
                            'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                            'st_id': result['student']['st_id'],
                            'score': float(result['score']),
                        })

                    second_year['term_two'] = sorted(second_year['term_two'], key=lambda x: float(x['score']), reverse=True)

                result_two_three = ResultsSerializer(
                    Result.objects.filter(
                        school=head.school,
                        academic_year=academic_year,
                        subject=subject_obj,
                        academic_term=3,
                        student_year=2,
                    ),
                    many=True).data
                if result_two_three:
                    for result in result_two_three:
                        second_year['term_three'].append({
                            'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                            'st_id': result['student']['st_id'],
                            'score': float(result['score']),
                        })

                    second_year['term_three'] = sorted(second_year['term_three'], key=lambda x: float(x['score']), reverse=True)

                    # Year Three Students Results
                    result_three_one = ResultsSerializer(
                        Result.objects.filter(
                            school=head.school,
                            academic_year=academic_year,
                            subject=subject_obj,
                            academic_term=1,
                            student_year=3,
                        ),
                        many=True).data
                    if result_three_one:
                        for result in result_three_one:
                            third_year['term_one'].append({
                                'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                                'st_id': result['student']['st_id'],
                                'score': float(result['score']),
                            })

                        third_year['term_one'] = sorted(third_year['term_one'], key=lambda x: float(x['score']),
                                                         reverse=True)

                    result_three_two = ResultsSerializer(
                        Result.objects.filter(
                            school=head.school,
                            academic_year=academic_year,
                            subject=subject_obj,
                            academic_term=2,
                            student_year=3,
                        ),
                        many=True).data
                    if result_three_two:
                        for result in result_three_two:
                            third_year['term_two'].append({
                                'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                                'st_id': result['student']['st_id'],
                                'score': float(result['score']),
                            })

                        third_year['term_two'] = sorted(third_year['term_two'], key=lambda x: float(x['score']),
                                                         reverse=True)

                    result_three_three = ResultsSerializer(
                        Result.objects.filter(
                            school=head.school,
                            academic_year=academic_year,
                            subject=subject_obj,
                            academic_term=3,
                            student_year=3,
                        ),
                        many=True).data
                    if result_three_three:
                        for result in result_three_three:
                            third_year['term_three'].append({
                                'name': f"{result['student']['user']['first_name']} {result['student']['user']['last_name']}",
                                'st_id': result['student']['st_id'],
                                'score': float(result['score']),
                            })

                        third_year['term_three'] = sorted(third_year['term_three'], key=lambda x: float(x['score']),
                                                           reverse=True)

            data.append({
                'department': department['name'],
                'year_one': first_year,
                'year_two': second_year,
                'year_three': third_year,
            })

        return Response(data)

    else:
        return Response(status=401)


# Admin
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def get_sch_admin_data(request):
    sch_admin = request.user.staff

    if request.method == 'GET':
        # Students Data
        classes = ClasseWithSubjectsSerializer(Classe.objects.filter(school=sch_admin.school, is_active=True), many=True).data
        year_one_unsort = [x for x in classes if x['students_year'] == 1]
        year_one = sorted(year_one_unsort, key=lambda x: x['created_at'], reverse=True)
        year_two_unsort = [x for x in classes if x['students_year'] == 2]
        year_two = sorted(year_two_unsort, key=lambda x: x['created_at'], reverse=True)
        year_three_unsort = [x for x in classes if x['students_year'] == 3]
        year_three = sorted(year_three_unsort, key=lambda x: x['created_at'], reverse=True)

        clas_items = [class_item for class_item in sorted(classes, key=lambda y: y['created_at'], reverse=True)]
        class_names_data = []
        for item in clas_items:
            class_names_data.append(f"{item['name']} FORM-{item['students_year']}")

        # Staff Data
        departments = DepartmentSerializer(Department.objects.filter(school=sch_admin.school), many=True).data
        programs = ProgramSerializer(Program.objects.filter(schools=sch_admin.school), many=True).data
        department_names = [x['name'] for x in departments]
        programs_names = [x['name'] for x in programs]
        academic_years = AcademicYearSerializer(AcademicYear.objects.filter(school=sch_admin.school).order_by('-start_date'), many=True).data
        subject_names = []
        for program in programs:
            for subject in program['subjects']:
                if subject['name'] not in subject_names:
                    subject_names.append(subject['name'])

        # Head Data
        heads_data = HeadSerializer(Head.objects.filter(school=sch_admin.school), many=True).data

        return Response({
            'classes': {
                'year_one': year_one,
                'year_two': year_two,
                'year_three': year_three
            },
            'class_names': class_names_data,
            'departments': departments,
            'department_names': department_names,
            'heads': heads_data,
            'subjects': subject_names,
            'programs': programs_names,
            'academic_years': academic_years,
        })

    else:
        data = request.data
        if data['type'] == 'create-year':
            current_year = AcademicYear.objects.get(school=sch_admin.school, name=data['year'])
            start_date = datetime.strptime(data['startDate'], "%Y-%m-%d").date()
            end_date = datetime.strptime(data['endDate'], "%Y-%m-%d").date()
            term_1_end_date = datetime.strptime(data['term2EndDate'], "%Y-%m-%d").date()
            term_2_start_date = datetime.strptime(data['term2StartDate'], "%Y-%m-%d").date()
            term_2_end_date = datetime.strptime(data['term2EndDate'], "%Y-%m-%d").date()
            current_date = timezone.now().date()

            duration = end_date - start_date

            if current_year.end_date >= current_date:
                return Response({
                    'ms': f"You cannot create a new academic year when the current academic year is still in progress"
                }, status=201)

            if current_year.end_date >= start_date:
                return Response({
                    'ms': f"The academic year start date must be greater than the current academic year end date"
                }, status=201)

            if duration.days < 6 * 30:
                return Response({
                    'ms': f"The duration between the academic year's start and end dates must be more than 6 months"
                }, status=201)

            try:
                AcademicYear.objects.get(school=sch_admin.school, name=data['yearName'])
                return Response({'ms': f"Academic year with name {data['yearName']} already exists"}, status=201)

            except AcademicYear.DoesNotExist:
                if start_date >= end_date:
                    return Response({
                        'ms': "The academic year end date must be greater than the start date"
                    }, status=201)

                elif term_1_end_date >= term_2_start_date:
                    return Response({
                        'ms': "Second semester/trimester start date must be greater than first semester/trimester end date"
                    }, status=201)

                elif term_2_start_date >= term_2_end_date:
                    return Response({
                        'ms': "Second semester/trimester end date must be greater than its start date"
                    }, status=201)

                elif not data['term3StartDate'] and term_2_end_date > end_date:
                    return Response({
                        'ms': "The academic year end date must be greater or equal to second semester end date"
                    }, status=201)

                if not data['term3StartDate'] and data['term3EndDate']:
                    return Response(
                        {'ms': "Third trimester start date cannot be empty when you specify its end date"},
                        status=201)

                if data['term3StartDate'] and data['term3StartDate'] != 'null':
                    if not data['term3EndDate']:
                        return Response(
                            {'ms': "Third trimester end date cannot be empty when you specify its start date"},
                            status=201)

                    term_3_start_date = datetime.strptime(data['term3StartDate'], "%Y-%m-%d").date()
                    term_3_end_date = datetime.strptime(data['term3EndDate'], "%Y-%m-%d").date()

                    if term_2_end_date >= term_3_start_date:
                        return Response({
                            'ms': "Third trimester start date must be greater than second trimester end date"
                        }, status=201)

                    elif term_3_start_date >= term_3_end_date:
                        return Response({'ms': "Third trimester end date must be greater than its end date"},
                                        status=201)

                    elif term_3_end_date > end_date:
                        return Response({
                            'ms': "The academic year end date must be greater or equal to trimester end date"
                        }, status=201)

                with transaction.atomic():
                    if data['term3StartDate'] and data['term3StartDate'] != 'null' and data['term3EndDate']:
                        academic_year = AcademicYear.objects.create(
                            name=data['yearName'],
                            school=sch_admin.school,
                            start_date=data['startDate'],
                            end_date=data['endDate'],
                            term_1_end_date=data['term2EndDate'],
                            term_2_start_date=data['term2StartDate'],
                            term_2_end_date=data['term2EndDate'],
                            term_3_start_date=data['term3StartDate'],
                            term_3_end_date=data['term3EndDate'],
                        )
                        academic_year.save()

                        new_academic_year = AcademicYear.objects.get(school=sch_admin.school, name=data['yearName'])
                        classes = Classe.objects.select_related('academic_year').prefetch_related('students').filter(
                            school=sch_admin.school,
                            is_active=True,
                        )

                        if classes.exists():
                            for clas in classes:
                                if clas.students_year == 3:
                                    clas.students_year = 4
                                    for student in clas.students.all():
                                        student.current_year = 4
                                        student.has_completed = True
                                        student.save()

                                    clas.is_active = False
                                    clas.save()

                                else:
                                    clas.students_year += 1
                                    clas.academic_years.add(new_academic_year)
                                    for student in clas.students.all():
                                        student.current_year += 1
                                        student.save()

                                    clas.save()

                        new_academic_data = AcademicYearSerializer(new_academic_year).data

                        return Response({
                            'ms': 'Academic year saved successfully',
                            'new_year': new_academic_data,
                        })

                    else:
                        academic_year = AcademicYear.objects.create(
                            name=data['yearName'],
                            school=sch_admin.school,
                            start_date=data['startDate'],
                            end_date=data['endDate'],
                            term_1_end_date=data['term2EndDate'],
                            term_2_start_date=data['term2StartDate'],
                            term_2_end_date=data['term2EndDate'],
                        )
                        academic_year.save()

                        new_academic_year = AcademicYear.objects.get(school=sch_admin.school, name=data['yearName'])
                        classes = Classe.objects.prefetch_related('students', 'academic_years').filter(
                            school=sch_admin.school,
                            is_active=True,
                        )

                        if classes.exists():
                            for clas in classes:
                                if clas.students_year == 3:
                                    clas.students_year = 4
                                    for student in clas.students.all():
                                        student.current_year = 4
                                        student.has_completed = True
                                        student.save()

                                    clas.is_active = False
                                    clas.save()

                                else:
                                    clas.students_year += 1
                                    clas.academic_years.add(new_academic_year)
                                    for student in clas.students.all():
                                        student.current_year += 1
                                        student.save()

                                    clas.save()

                        new_academic_data = AcademicYearSerializer(new_academic_year).data

                        return Response({
                            'ms': 'Academic year saved successfully',
                            'new_year': new_academic_data,
                        })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_staff(request):
    data = request.data
    sch_admin = request.user.staff
    if data['type'] == 'input-staff':
        department_name = data['departmentName']
        department = Department.objects.get(school=sch_admin.school, name=department_name)
        department_data = DepartmentSerializer(department).data

        user_no = random.randint(100, 999)
        username = f"{data['firstName'].replace(' ', '')[0].upper()}{data['lastName'].replace(' ', '').lower()}{user_no}"

        if Staff.objects.filter(school=sch_admin.school, staff_id=data['staffId']).exists():
            return Response({'ms': f"Staff with ID [ {data['staffId']} ] already exists"}, status=201)

        with transaction.atomic():
            try:
                user = User.objects.create_user(
                    username=username,
                    password=username,
                    first_name=data['firstName'].upper(),
                    last_name=data['lastName'].upper(),
                )
                user.save()

            except IntegrityError:
                return Response({'ms': 'A user with these details already exists'}, status=201)

            staff_user = User.objects.get(username=username)

            stf_obj = Staff.objects.create(
                user=staff_user,
                staff_id=data['staffId'],
                department=department,
                role='teacher',
                gender=data['gender'].upper(),
                dob=data['dob'].upper(),
                school=sch_admin.school
            )

            try:
                if isinstance(data['subjects'], str):
                    subjects = data['subjects'].split(',')
                    print(subjects)
                    for subject in subjects:
                        if subject != '':
                            subject_obj = Subject.objects.get(schools=sch_admin.school, name=subject)
                            try:
                                if subject_obj in department.subjects.all():
                                    stf_obj.subjects.add(subject_obj)
                                else:
                                    raise SuspiciousOperation

                            except SuspiciousOperation:
                                transaction.set_rollback(True)
                                return Response({'ms': f"The {department_name} department does not teach {subject}"},
                                                status=201)

                elif isinstance(data['subjects'], list):
                    for subject in data['subjects']:
                        subject_obj = Subject.objects.get(name=subject)
                        try:
                            if subject_obj in department.subjects:
                                stf_obj.subjects.add(subject_obj)
                            else:
                                raise SuspiciousOperation

                        except SuspiciousOperation:
                            transaction.set_rollback(True)
                            return Response({'ms': f"The {department_name} department does not teach {subject}"},
                                            status=201)

                stf_obj.save()

                staff_obj = Staff.objects.get(staff_id=data['staffId'])
                department.teachers.add(staff_obj)
                department.save()

                response_staff = SpecificStaffSerializer(staff_obj).data
                return Response({
                    'ms': 'Teacher successfully created and saved',
                    'staff': response_staff,
                    'department_name': department_data['name'],
                })

            except IntegrityError:
                return Response({'ms': 'Staff with these details already exists'}, status=201)

    elif data['type'] == 'get-staff-file':
        sch_admin_data = StaffSerializer(sch_admin).data
        department_name = data['departmentName']
        if Department.objects.filter(school=sch_admin.school, name=department_name).exists():
            wb = load_workbook('staticfiles/files/create_staff.xlsx')
            ws = wb.worksheets[0]
            filename = f"{department_name}.xlsx"
            ws.title = department_name
            byte_file = io.BytesIO()
            wb.save(byte_file)

            if settings.DEBUG:
                file_path = f"{get_school_folder(sch_admin_data['school']['name'])}/staff/{sch_admin_data['user']['username']}/{filename}"
                if default_storage.exists(file_path):
                    default_storage.delete(file_path)

                save_file = default_storage.save(file_path, byte_file)

                return Response({
                    'filename': f"{filename}.xlsx",
                    'file_path': f"http://localhost:8000{default_storage.url(save_file)}",
                })

            else:
                file_path = f"media/{get_school_folder(sch_admin_data['school']['name'])}/staff/{sch_admin_data['user']['username']}/{filename}"
                if default_storage.exists(file_path):
                    default_storage.delete(file_path)

                save_file = default_storage.save(file_path, byte_file)

                return Response({
                    'filename': f"{filename}.xlsx",
                    'file_path': default_storage.url(save_file),
                })

    elif data['type'] == 'upload-staff-file':
        department_name = data['departmentName']
        wb = load_workbook(data['file'])
        if department_name == wb.sheetnames[0]:
            wb.close()
            df = pd.read_excel(data['file'], sheet_name=department_name)
            df_data = df.iloc[3:, :6].dropna().values.tolist()

            existing_staff = Staff.objects.filter(school=sch_admin.school,
                                                  staff_id__in=[stf[2] for stf in df_data]).first()
            if existing_staff:
                return Response({
                    'ms': f"Staff with ID [ {SpecificStaffSerializer(existing_staff).data['staff_id']} ] already exists"},
                    status=201)

            with transaction.atomic():
                staff_ids = []
                department = Department.objects.get(school=sch_admin.school, name=department_name)
                user_no = random.randint(100, 999)

                for stf in df_data:
                    username = f"{stf[0].replace(' ', '')[0].upper()}{stf[1].replace(' ', '').lower()}{user_no}"
                    user = User.objects.create_user(
                        username=username,
                        password=username,
                        first_name=stf[0].upper(),
                        last_name=stf[1].upper(),
                    )
                    user.save()

                    staff_user = User.objects.get(username=username)

                    if str(stf[3]).upper() == 'MALE' or str(stf[3]).upper() == 'FEMALE':
                        pass

                    else:
                        transaction.set_rollback(True)
                        return Response({
                            'ms': f"{stf[3]} is not a valid gender. Valid values are 'MALE' or 'FEMALE'. Recheck the gender for {stf[0]} {stf[1]} in the file"},
                            status=201)

                    try:
                        stf_obj = Staff.objects.create(
                            user=staff_user,
                            staff_id=stf[2],
                            role='teacher',
                            department=department,
                            gender=stf[3].upper(),
                            dob=stf[4],
                            school=sch_admin.school
                        )

                    except IntergrityError_unique_constraint:
                        return Response({
                            'ms': f"{stf[0]} {stf[1]}'s staff ID matches with another teacher's staff ID. Recheck that the ID {stf[2]} is unique in the file"
                        }, status=201)

                    except ValidationError:
                        transaction.set_rollback(True)
                        return Response({
                            'ms': f"{stf[4]} is not a valid date, Recheck the date of birth of {stf[0]} {stf[1]} in the file"},
                            status=201)

                    subjects = stf[5].split(',')
                    for subject in subjects:
                        try:
                            subject_obj = Subject.objects.get(schools=sch_admin.school, name=subject.strip().upper())
                            if subject_obj in department.subjects.all():
                                stf_obj.subjects.add(subject_obj)
                            else:
                                raise SuspiciousOperation

                        except Subject.DoesNotExist:
                            transaction.set_rollback(True)
                            return Response({
                                'ms': f"{subject} is not a valid subject name. Recheck the subject(s) for {stf[0]} {stf[1]} in the file"
                            }, status=201)

                        except SuspiciousOperation:
                            transaction.set_rollback(True)
                            return Response({
                                'ms': f"The {department_name} department does not teach {subject}. Recheck the subject(s) of {stf[0]} {stf[1]}"
                            }, status=201)

                    stf_obj.save()

                    staff_ids.append(stf[2])

                for staff_id in staff_ids:
                    staff_obj = Staff.objects.get(school=sch_admin.school, staff_id=staff_id)
                    department.teachers.add(staff_obj)

                department.save()

            department_data = DepartmentSerializer(Department.objects.get(school=sch_admin.school, name=department_name)).data
            return Response({
                'ms': 'File uploaded and teachers saved successfully',
                'data': department_data,
            })

        else:
            return Response({'ms': 'The first sheet name must be the same as the selected department'}, status=201)

    elif data['type'] == 'delete-staff':
        staff_data = SpecificStaffSerializer(Staff.objects.get(school=sch_admin.school, staff_id=data['staffId'])).data
        user = User.objects.get(username=staff_data['user']['username'])
        with transaction.atomic():
            # delete user object
            user.delete()
            teacher = Staff.objects.get(staff_id=data['staffId'])
            department = Department.objects.get(school=sch_admin.school, teachers=teacher)
            # remove staff from department
            department.teachers.remove(teacher)
            department.save()
            # delete staff object
            teacher.delete()

        return Response({
            'department_name': DepartmentSerializer(department).data['name'],
        })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_head(request):
    data = request.data
    sch_admin = request.user.staff
    if data['type'] == 'create-head':
        user_no = random.randint(100, 999)
        username = f"{data['firstName'].replace(' ', '')[0].upper()}{data['lastName'].replace(' ', '').lower()}{user_no}"

        if Head.objects.filter(school=sch_admin.school, head_id=data['headId']).exists():
            return Response({'ms': f"Head with ID [ {data['headId']} ] already exists"}, status=201)

        with transaction.atomic():
            try:
                user = User.objects.create_user(
                    username=username,
                    password=username,
                    first_name=data['firstName'].upper(),
                    last_name=data['lastName'].upper(),
                )
                user.save()

            except IntegrityError:
                return Response({'ms': 'A user with these details already exists'}, status=201)

            head_user = User.objects.get(username=username)
            head_obj = Head.objects.create(
                user=head_user,
                head_id=data['headId'],
                role=data['role'].lower().replace(' ', '_'),
                gender=data['gender'].upper(),
                dob=data['dob'].upper(),
                school=sch_admin.school
            )

            try:
                head_obj.save()

                new_head = Head.objects.get(head_id=data['headId'])

                response_head = HeadSerializer(new_head).data
                return Response(response_head)

            except IntegrityError:
                return Response({'ms': 'Head with these details already exists'}, status=201)

    elif data['type'] == 'delete-head':
        head = Head.objects.select_related('user').get(school=sch_admin.school, head_id=data['headId'])
        user = User.objects.get(username=head.user.username)
        with transaction.atomic():
            # delete user object
            user.delete()
            # delete head object
            head.delete()

        return Response(status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_students(request):
    sch_admin = request.user.staff
    data = request.data
    if data['type'] == 'create-class':
        program = Program.objects.get(name=data['program'])
        academic_year = AcademicYear.objects.get(school=sch_admin.school, name=data['year'])
        class_name = f"{data['year'].split('-')[0]}-{data['className'].replace(' ', '-')}"
        program_data = ProgramSerializer(program).data
        program_subjects = [sub['name'] for sub in program_data['subjects']]
        subjects = data['subjects'].split(',')

        try:
            Classe.objects.get(school=sch_admin.school, name=class_name.upper())
            return Response({'ms': 'Class with this name already exists'}, status=201)

        except Classe.DoesNotExist:
            for subject in subjects:
                if subject not in program_subjects:
                    return Response({'ms': f"Students enrolled in {data['program']} do not study {subject}"},
                                    status=201)

            with transaction.atomic():
                clas = Classe.objects.create(
                    school=sch_admin.school,
                    program=program,
                    name=class_name.upper(),
                    students_year=1,
                    date_enrolled=data['enrollmentDate'],
                    completion_date=data['completionDate']
                )

                clas.academic_years.add(academic_year)
                for subject in subjects:
                    if subject != '':
                        subject_obj = Subject.objects.get(schools=sch_admin.school, name=subject)
                        clas.subjects.add(subject_obj)

                clas.save()

                new_class = ClasseWithSubjectsSerializer(Classe.objects.get(school=sch_admin.school, name=class_name.upper())).data
                return Response({
                    'ms': "Class successfully created",
                    'new_class': new_class,
                    'class_name': class_name.upper()
                })

    if data['type'] == 'input-student':
        class_name = data['className'].split()[0]
        clas = Classe.objects.select_related('program').get(school=sch_admin.school, name=class_name)
        class_data = ClasseSerializer(clas).data

        user_no = random.randint(100, 999)
        username = f"{data['firstName'].replace(' ', '')[0].upper()}{data['lastName'].replace(' ', '').lower()}{user_no}"

        existing_student_ids = set(Student.objects.filter(
            school=sch_admin.school,
            current_year=1
        ).values_list('st_id', flat=True))

        st_id = f"{data['year'].split('-')[0]}{random.randint(1000,3000)}"
        while st_id in existing_student_ids:
            st_id = f"{data['year'].split('-')[0]}{random.randint(1000,3000)}"

        with transaction.atomic():
            try:
                user = User.objects.create_user(
                    username=username,
                    password=username,
                    first_name=data['firstName'].upper(),
                    last_name=data['lastName'].upper(),
                )
                user.save()

            except IntegrityError:
                return Response({'ms': 'A user with these details already exists'}, status=201)

            st_user = User.objects.get(username=username)

            student = Student.objects.create(
                user=st_user,
                program=clas.program,
                st_class=clas,
                current_year=class_data['students_year'],
                date_enrolled=class_data['date_enrolled'],
                st_id=st_id,
                gender=data['gender'].upper(),
                dob=data['dob'].upper(),
                school=sch_admin.school
            )

            try:
                student.save()
                st_obj = Student.objects.get(school=sch_admin.school, st_id=st_id)
                clas.students.add(st_obj)
                clas.save()

                response_student = SpecificStudentSerializer(st_obj).data
                return Response({
                    'ms': 'student successfully created and saved',
                    'student': response_student,
                    'year': 'yearOne' if class_data['students_year'] == 1 else ('yearTwo' if class_data['students_year'] == 2 else 'yearThree'),
                    'class_name': class_data['name'],
                })

            except IntegrityError:
                return Response({'ms': 'student with these details already exists'}, status=201)

    elif data['type'] == 'get-students-file':
        sch_admin_data = StaffSerializer(sch_admin).data
        class_name = data['className'].split()[0]
        clas = Classe.objects.filter(school=sch_admin.school, name=class_name).first()
        if clas:
            clas_data = ClasseSerializer(clas).data
            wb = load_workbook('staticfiles/files/students_admission.xlsx')
            ws = wb.worksheets[0]
            filename = f"{clas_data['name']}-FORM-{clas_data['students_year']}.xlsx"
            ws.title = f"{clas_data['name']} FORM-{clas_data['students_year']}"
            byte_file = io.BytesIO()
            wb.save(byte_file)

            if settings.DEBUG:
                file_path = f"{get_school_folder(sch_admin_data['school']['name'])}/staff/{sch_admin_data['user']['username']}/{filename}"
                if default_storage.exists(file_path):
                    default_storage.delete(file_path)

                save_file = default_storage.save(file_path, byte_file)

                return Response({
                    'filename': f"{filename}.xlsx",
                    'file_path': f"http://localhost:8000{default_storage.url(save_file)}",
                })

            else:
                file_path = f"media/{get_school_folder(sch_admin_data['school']['name'])}/staff/{sch_admin_data['user']['username']}/{filename}"
                if default_storage.exists(file_path):
                    default_storage.delete(file_path)

                save_file = default_storage.save(file_path, byte_file)

                return Response({
                    'filename': f"{filename}.xlsx",
                    'file_path': default_storage.url(save_file),
                })

    elif data['type'] == 'upload-students-file':
        class_name = data['className'].split()[0]
        wb = load_workbook(data['file'])
        if data['className'] == wb.sheetnames[0]:
            wb.close()
            df = pd.read_excel(data['file'], sheet_name=data['className'])
            df_data = df.iloc[2:, :4].dropna().values.tolist()

            existing_student_ids = set(Student.objects.filter(
                school=sch_admin.school,
                current_year=1
            ).values_list('st_id', flat=True))

            students_id = set()
            while len(students_id) < len(df_data):
                new_id = f"{data['year'].split('-')[0]}{random.randint(1000, 3000)}"
                if new_id not in students_id and new_id not in existing_student_ids:
                    students_id.add(new_id)

            students_id = list(students_id)

            with transaction.atomic():
                clas = Classe.objects.select_related('program').get(school=sch_admin.school, name=class_name)
                user_no = random.randint(100, 999)

                for st, st_id in zip(df_data, students_id):
                    username = f"{st[0].replace(' ', '')[0].upper()}{st[1].replace(' ', '').lower()}{user_no}"
                    user = User.objects.create_user(
                        username=username,
                        password=username,
                        first_name=st[0].upper(),
                        last_name=st[1].upper(),
                    )
                    user.save()

                    st_user = User.objects.get(username=username)

                    try:
                        student = Student.objects.create(
                            user=st_user,
                            program=clas.program,
                            st_id=st_id,
                            current_year=clas.students_year,
                            st_class=clas,
                            date_enrolled=clas.date_enrolled,
                            gender=st[2].upper(),
                            dob=st[3],
                            school=sch_admin.school,
                        )
                        student.save()

                    except ValidationError:
                        transaction.set_rollback(True)
                        return Response({
                            'ms': f"{st[3]} is not a valid date, recheck the date of birth of {st[0]} {st[1]} in the file"},
                            status=201)

                for st_id in students_id:
                    st_obj = Student.objects.get(school=sch_admin.school, st_id=st_id)
                    clas.students.add(st_obj)

                clas.save()

            class_data = ClasseSerializer(Classe.objects.get(school=sch_admin.school, name=class_name)).data
            return Response({
                'ms': 'File uploaded and students saved successfully',
                'data': class_data,
                'year': 'yearOne' if class_data['students_year'] == 1 else (
                    'yearTwo' if class_data['students_year'] == 2 else 'yearThree'),
            })

        else:
            return Response({'ms': 'The first sheet name must be same as the selected class'}, status=201)

    elif data['type'] == 'delete-student':
        st_data = SpecificStudentSerializer(Student.objects.get(st_id=data['stId'])).data
        class_data = ClasseSerializer(Classe.objects.get(name=data['className'])).data
        user = User.objects.get(username=st_data['user']['username'])

        with transaction.atomic():
            # delete user object
            user.delete()
            student = Student.objects.get(st_id=data['stId'])
            clas = Classe.objects.get(name=data['className'])
            # remove student from class
            clas.students.remove(student)
            clas.save()
            # delete student object
            student.delete()

        return Response({
            'class_name': class_data['name'],
            'year': 'yearOne' if class_data['students_year'] == 1 else (
                'yearTwo' if class_data['students_year'] == 2 else 'yearThree'),
        })

    elif data['type'] == 'delete-subject':
        subject = Subject.objects.get(name=data['subject'])
        clas = Classe.objects.get(school=sch_admin.school, name=data['className'])
        class_data = ClasseWithoutStudentsSerializer(clas).data

        with transaction.atomic():
            # remove subject from class
            clas.subjects.remove(subject)
            clas.save()

        return Response({
            'class_name': class_data['name'],
            'year': 'yearOne' if class_data['students_year'] == 1 else (
                'yearTwo' if class_data['students_year'] == 2 else 'yearThree'),
        })

    elif data['type'] == 'delete-class':
        clas = Classe.objects.get(school=sch_admin.school, name=data['className'])

        with transaction.atomic():
            # remove subject from class
            clas.delete()

        return Response(status=200)


# TEACHER, HOD, HEAD
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def staff_notification(request):
    if request.method == 'POST' and request.data['type'] == 'send':
        try:
            sender = request.user.staff
            staff_ids = json.loads(request.POST.get('staff_ids'))
            if staff_ids and len(staff_ids) > 0:
                staff_message_instance = StaffNotification.objects.create(
                    content=request.data['message'],
                    sent_by_hod=sender
                )
                for staff_id in staff_ids:
                    send_to = Staff.objects.get(school=sender.school, staff_id=staff_id)
                    staff_message_instance.send_to.add(send_to)

                staff_message_instance.save()

                messages = get_staff_messages(request.user)

                return Response(messages)
            else:
                return Response(status=401)

        except User.staff.RelatedObjectDoesNotExist:
            sender = request.user.head
            staff_ids = json.loads(request.POST.get('staff_ids'))
            if staff_ids and len(staff_ids) > 0:
                head_message_instance = StaffNotification.objects.create(
                    content=request.data['message'],
                    sent_by_head=sender
                )
                for stf_id in staff_ids:
                    send_to = Staff.objects.get(school=sender.school, staff_id=stf_id)
                    head_message_instance.send_to.add(send_to)

                head_message_instance.save()

                messages = get_staff_messages(request.user)

                return Response(messages)
            else:
                return Response(status=401)

    elif request.method == 'POST' and request.data['type'] == 'delete':
        message_id = int(request.data['id'])
        try:
            hod = request.user.staff
            hod_messages = get_object_or_404(StaffNotification, sent_by_hod=hod, id=message_id)
            if hod_messages:
                hod_messages.delete()

            messages = get_staff_messages(request.user)

            return Response(messages)

        except User.staff.RelatedObjectDoesNotExist:
            head = request.user.head
            head_messages = get_object_or_404(StaffNotification, sent_by_head=head, id=message_id)
            if head_messages:
                head_messages.delete()

            messages = get_staff_messages(request.user)

            return Response(messages)

    else:
        try:
            head = request.user.head
            head_staff_data = []
            staff = SpecificStaffSerializer(Staff.objects.filter(school=head.school), many=True).data
            for stf in staff:
                head_staff_data.append({
                    'name': f"{stf['user']['first_name']} {stf['user']['last_name']}",
                    'staff_id': stf['staff_id']
                })

            messages = get_staff_messages(request.user)

            return Response({'messages': messages, 'head_staff_data': head_staff_data})

        except User.head.RelatedObjectDoesNotExist:
            messages = get_staff_messages(request.user)

            return Response(messages)


# OTHER
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def user_help(request):
    problem = request.data['problem']
    subject = f"School System"
    try:
        student = request.user.student
        st = StudentSerializer(student).data
        subject = f"{st['school']['name']} (student) [ Name: {st['user']['first_name']} {st['user']['last_name']} Username: {st['user']['username']} ]"
    except User.student.RelatedObjectDoesNotExist:
        staff = request.user.staff
        stf = StaffSerializer(staff).data
        subject = f"{stf['school']['name']} ({stf['role']}) [ Name: {stf['user']['first_name']} {stf['user']['last_name']} Username: {stf['user']['username']} ]"
    except User.staff.RelatedObjectDoesNotExist:
        head = request.user.head
        hd = HeadSerializer(head).data
        subject = f"{hd['school']['name']} ({hd['role']}) [ Name: {hd['user']['first_name']} {hd['user']['last_name']} Username: {hd['user']['username']}]"

    try:
        send_mail(subject, problem, 'nyamejustice2000@gmail.com', ['nyamejustice2000@gmail.com'], fail_silently=False)
        return Response(status=200)

    except Exception as e:
        return Response(status=401)


# class StudentsView(ListAPIView):
#     queryset = Student.objects.all()
#     serializer_class = StudentSerializer
#

# class StaffView(ListAPIView):
#     queryset = Staff.objects.all()
#     serializer_class = StaffSerializer

#
# class ClasseView(ListAPIView):
#     queryset = Classe.objects.all()
#     serializer_class = ClasseSerializer
#
#
# class SubjectAssignmentView(ListAPIView):
#     queryset = SubjectAssignment.objects.all()
#     serializer_class = SubjectAssignmentSerializer
#
#
class FileView(ListAPIView):
    queryset = File.objects.all()
    serializer_class = FileSerializer


def root(request):
    return HttpResponse("<h1>EduAAP/h1>")
