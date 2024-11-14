# Django
from django.contrib.auth.models import User
from django.db import IntegrityError, transaction
from django.db.utils import IntegrityError as IntergrityError_unique_constraint
from django.core.files.storage import default_storage
from django.core.exceptions import SuspiciousOperation
from django.core.exceptions import ValidationError
from django.core.validators import EmailValidator
from django.utils import timezone
from django.http import FileResponse
from django.forms import DateField


# Document Manipulation
import pandas as pd
from openpyxl.styles import Font, Alignment, Protection
from openpyxl import Workbook, load_workbook

from api.models import *
from api.serializer import *
from collections import defaultdict
import random
import io
from datetime import datetime
import time
import json
from api.utils import format_staff_creation_file, ErrorMessageException, valid_phone_number, validate_nationality, validate_country_region, get_country_from_nationality, get_country_regions

# Django Restframework
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated


# Admin
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def school_admin_data(request):
    sch_admin = request.user.staff
    school = sch_admin.school

    if request.method == 'GET':
        academic_years = AcademicYearSerializer(AcademicYear.objects.filter(school=school).order_by('-start_date'), many=True).data
        student_classes = Classe.objects.filter(school=school).order_by('students_year')
        student_classes_data = ClassesSerializerOne(student_classes, many=True).data
        linked_classes = AdminLinkedClassSerializer(LinkedClasse.objects.filter(school=school), many=True).data
        departments = []
        programs = []
        if school.level.name == 'SHS':
            departments_data = DepartmentNameSerializer(Department.objects.filter(school=school), many=True).data
            departments = [_department['name'] for _department in departments_data]
            programs_data = ProgramNameSerializer(Program.objects.filter(schools=school), many=True).data
            programs = [_program['name'] for _program in programs_data]
            
        staff = StaffSerializerOne(Staff.objects.filter(school=school, is_active=True).order_by('-date_created'), many=True).data
        subjects = SubjectsSerializer(Subject.objects.filter(schools=school), many=True).data
        subject_names = [_subject['name'] for _subject in subjects]
        
        return Response({
            'classes': student_classes_data,
            #     'year_one': year_one,
            #     'year_two': year_two,
            #     'year_three': year_three
            # },
            # 'class_names': class_names_data,
            'departments': departments,
            # 'department_names': department_names,
            # 'heads': heads_data,
            'subjects': subject_names,
            'staff': staff,
            'programs': programs,
            'academic_years': academic_years,
            'linked_classes': linked_classes,
        }, status=200)
    
    elif request.method == 'POST':  
        data = request.data
        if data['type'] == 'assignClassHeadTeacher':
            staff_id = data['teacherId']
            new_head_teacher = Staff.objects.get(school=school, staff_id=staff_id)
            students_class = Classe.objects.get(school=school, name=data['studentsClass'])
            with transaction.atomic():
                try:
                    students_class.head_teacher = new_head_teacher
                    students_class.save()
                except Exception as e:
                    transaction.set_rollback(True)
                    return Response(status=400)
            
            new_head_teacher_data = StaffSerializerThree(new_head_teacher).data
            return Response(new_head_teacher_data, status=200)
        

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def school_admin_linked_class(request):
    sch_admin = request.user.staff
    school = sch_admin.school
    data = request.data
    from_class = Classe.objects.select_related('program').get(school=school, name=data['fromClass'])
    to_class = Classe.objects.select_related('program').get(school=school, name=data['toClass'])
    
    if data['type'] == 'create':
        if from_class.students_year >= to_class.students_year:  
            return Response({'message': f"The class the student will be promoted to must be above this class, {from_class.name}"}, status=400)
        
        if (from_class.program and not to_class.program) or (from_class.program and to_class.program != from_class.program):
            return Response({'message': "The two classes must be under the same program"}, status=400)
        
        existing_linked_class = LinkedClasse.objects.filter(school=school, from_class=from_class).exists()
        if existing_linked_class:
            return Response({'message': f"You have already linked a class to {from_class.name}"}, status=400)
        
        linked_class = LinkedClasse.objects.create(
            school=school,
            from_class=from_class,
            to_class=to_class,
        )
        with transaction.atomic():
            try:
                linked_class.save()
            except IntegrityError:
                transaction.set_rollback(True)
                return Response({'message': f'You have already linked a class to {to_class.name}'}, status=400)
            except Exception as e:
                transaction.set_rollback(True)
                return Response(status=400)
        
        new_linked_class_data = AdminLinkedClassSerializer(LinkedClasse.objects.get(school=school, from_class=from_class, to_class=to_class)).data
        return Response(new_linked_class_data, status=200)
    
    elif data['type'] == 'delete':
        linked_class = LinkedClasse.objects.get(school=school, from_class=from_class, to_class=to_class)
        with transaction.atomic():
            try:
                linked_class.delete()
            except Exception as e:
                transaction.set_rollback(True)
                return Response(status=400)
        
        return Response(status=200)
    
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def school_admin_academic_years(request):
    sch_admin = request.user.staff
    school = sch_admin.school
    data = request.data
    
    def parse_date(date:str):
        return datetime.strptime(date, "%Y-%m-%d").date()
                
    if data['type'] == 'create':
        new_year_name = data['newYear']
        current_year = AcademicYear.objects.get(school=school, name=data['year'])
        start_date = parse_date(data['startDate'])
        end_date = parse_date(data['endDate'])
        students_graduation_date = parse_date(data['graduationDate']) if data['graduationDate'] else None 
        repeated_students_ids = json.loads(data['repeatedStudents'])
        period_division = int(data['periodDivision'])
        term_1_end_date = parse_date(data['term1EndDate'])
        term_2_start_date = parse_date(data['term2StartDate'])
        term_2_end_date = parse_date(data['term2EndDate'])
        term_3_start_date = None if period_division == 2 else parse_date(data['term3StartDate'])
        term_3_end_date = None if period_division == 2 else parse_date(data['term3EndDate'])
        period_division_name = data['periodDivisionName']
        current_date = timezone.now().date()
        duration = end_date - start_date
        
        def validate_dates():
            validation_error = None
            if current_year.end_date >= current_date:
                validation_error = "You cannot create a new academic year when the current academic year is still in progress"
        
            elif current_year.end_date >= start_date:
                validation_error = f"The academic year start date must be greater than the current academic year[ {current_year.name} ] end date"

            elif duration.days < 6 * 30:
                validation_error = "The duration between the academic year's start and end dates must be more than 6 months"

            elif start_date >= end_date:
                validation_error = "The academic year end date must be greater than the start date"

            elif term_1_end_date >= term_2_start_date:
                validation_error = f"{period_division_name} 2 start date must be greater than {period_division_name} 1 end date"

            elif term_2_start_date >= term_2_end_date:
                validation_error = f"{period_division_name} 2 end date must be greater than its start date"

            elif period_division == 2 and term_2_end_date > end_date:
                validation_error = f"The academic year end date must be greater or equal to {period_division} 2 end date"
                    
            elif period_division == 3:
                if term_2_end_date >= term_3_start_date:
                    validation_error = f"{period_division} 3 start date must be greater than {period_division} 2 end date"

                elif term_3_start_date >= term_3_end_date:
                    validation_error = f"{period_division} 3 end date must be greater than its start date"

                elif term_3_end_date > end_date:
                    validation_error = f"The academic year end date must be greater or equal to {period_division} 3 end date"
                    
            return validation_error
    
        def create_new_academic_year():
            academic_year = AcademicYear.objects.create(
                name=new_year_name,
                school=school,
                start_date=start_date,
                end_date=end_date,
                term_1_end_date=term_1_end_date,
                term_2_start_date=term_2_start_date,
                term_2_end_date=term_2_end_date,
                term_3_start_date=term_3_start_date,
                term_3_end_date=term_3_end_date,
                students_graduation_date=students_graduation_date,
            )
            academic_year.save()
            new_academic_year = AcademicYear.objects.get(school=school, name=new_year_name)
            return new_academic_year
        
        invalid_validation = validate_dates()
        if invalid_validation:
            return Response({'message': invalid_validation}, status=400)
        try:
            AcademicYear.objects.get(school=school, name=new_year_name)
            return Response({'message': f"Academic year with name {new_year_name} already exists"}, status=400)
        except AcademicYear.DoesNotExist:
            with transaction.atomic():
                try:
                    repeated_students = Student.objects.filter(school=school, st_id__in=repeated_students_ids).distint()
                    all_repeated_students = []
                    for _student in repeated_students:
                        _student.repeated = True
                        all_repeated_students.append(_student)
                        
                    Student.objects.bulk_update(all_repeated_students, ['repeated'])
                    new_academic_year = create_new_academic_year()
                    classes = Classe.objects.prefetch_related('students').filter(school=school)
                    old_classes_to_update = []
                    new_classes_to_update = []
                    graduation_classes_to_create = []
                    graduated_students = []
                    promoted_students = []
                    repeated_students_to_update = []
                    years_to_complete = school.level.years_to_complete
                    for _class in classes:
                        try:
                            students = _class.students.all()
                            if _class.students_year == years_to_complete:
                                graduation_class = GraduatedClasse.objects.create(
                                    school=school,
                                    name=f"{_class.name} CLASS OF {current_year.name}",
                                    graduated_class_name=_class.name,
                                    graduation_year=current_year,
                                    students_year=years_to_complete,
                                    head_teacher=_class.head_teacher,
                                    program=_class.program,
                                )
                                graduation_class.subjects.add(_class.subjects.all())
                                for _student in students:
                                    if _student.repeated:
                                        _student.repeated_academic_years.add(current_year)
                                        _student.academic_years.add(new_academic_year)
                                        _student.repeated = False
                                        repeated_students_to_update.append(_student)
                                    else:
                                        _student.current_year += 1
                                        _student.has_completed = True
                                        _student.graduation_date = current_year.students_graduation_date
                                        _student.graduation_year = current_year
                                        _class.students.remove(_student)
                                        graduated_students.append(_student)
                                        graduation_class.students.add(_student)
                                graduation_classes_to_create.append(graduation_class)
                            else:
                                linked_class = LinkedClasse.objects.prefetch_related('to_class__students').get(school=school, from_class=_class)
                                new_students_class = linked_class.to_class
                                for _student in students:
                                    if _student.repeated:
                                        _student.repeated_academic_years.add(current_year)
                                        _student.repeated = False
                                        repeated_students_to_update.append(_student)
                                    else:
                                        _student.current_year += 1
                                        _student.linked_classes.add(linked_class)
                                        _student.academic_years.add(new_academic_year)
                                        _student.st_class = new_students_class
                                        new_students_class.students.add(_student)
                                        _class.students.remove(_student)
                                        promoted_students.append(_student)
                                new_classes_to_update.append(new_students_class)
                            old_classes_to_update.append(_class)
                        except LinkedClasse.DoesNotExist:
                            transaction.set_rollback(True)
                            return Response({'message': f"No class has been linked to the {_class.name} class"}, status=400)
                        except Exception as e:
                            transaction.set_rollback(True)
                            print(e)
                            return Response(status=400)
                    
                    Student.objects.bulk_update(graduated_students, ['current_year', 'has_completed', 'graduation_year', 'graduation_date'])
                    Student.objects.bulk_update(promoted_students, ['current_year', 'academic_years', 'st_class', 'linked_classes'])
                    Student.objects.bulk_update(repeated_students_to_update, ['repeated', 'repeated_academic_years'])
                    Classe.objects.bulk_update(new_classes_to_update, ['students_year', 'academic_years'])
                    Classe.objects.bulk_update(old_classes_to_update, ['students_year', 'academic_years'])
                    GraduatedClasse.objects.bulk_create(graduation_class)
                except Exception as e:
                    transaction.set_rollback(True)
                    print(e)
                    return Response(status=400)
                    
            new_academic_year_data = AcademicYearSerializer(new_academic_year).data
            return Response({
                'message': 'Operation successful',
                'new_year': new_academic_year_data,
            })

    elif data['type'] == 'delete':
        all_academic_years = AcademicYear.objects.filter(school=school).order_by('-end_date')
        current_year = all_academic_years[0]
        previous_year = all_academic_years[1]
        existing_assignments = SubjectAssignment.objects.filter(school=school, academic_year=current_year).exists()
        if existing_assignments:
            return Response({'message': "There are teachers coursework in this academic year"}, status=400)

        with transaction.atomic():
            try:
                classes = Classe.objects.prefetch_related('students').filter(school=school)
                old_classes_to_update = []
                new_classes_to_update = []
                graduated_students_old_classes = []
                graduated_classes_to_delete = []
                undo_graduated_students = []
                undo_promoted_students = []
                undo_repeated_students = []
                years_to_complete = school.level.years_to_complete
                for _class in classes:
                    try:
                        if _class.students_year == years_to_complete:
                            graduated_class = GraduatedClasse.objects.prefetch_related('students', 'subjects').select_related('head_teacher', 'program').get(school=school, graduated_class_name=_class.name, graduation_year=previous_year)
                            students = graduated_class.students.all()
                            for _student in students:
                                _student.current_year -= 1
                                _student.st_class = _class
                                _student.graduation_year = None
                                _student.graduation_date = None
                                _student.has_completed = False
                                undo_graduated_students.append(_student)
                            _class.subjects.set(graduated_class.subjects.all())
                            _class.head_teacher = graduated_class.head_teacher
                            _class.program = graduated_class.program
                            _class.name = graduated_class.graduated_class_name
                            _class.students_year = graduated_class.students_year
                            _class.students.set(students)
                            graduated_students_old_classes.append(_class)
                            graduated_classes_to_delete.append(graduated_class)
                        else:
                            linked_class = LinkedClasse.prefetch_related('to_class__students').objects.get(school=school, from_class=_class)
                            promoted_students_class = linked_class.to_class
                            promoted_students = promoted_students_class.students.all()
                            for _student in promoted_students:
                                if current_year in _student.academic_years.all() and previous_year not in _student.repeated_academic_years.all():
                                    _student.current_year -= 1
                                    _student.linked_classes.remove(linked_class)
                                    _student.academic_years.remove(current_year)
                                    _student.st_class = _class
                                    promoted_students_class.students.remove(_student)
                                    _class.students.add(_student)
                                    undo_promoted_students.append(_student)
                                elif current_year in _student.academic_years.all() and previous_year in _student.repeated_academic_years.all():
                                    _student.repeated_academic_years.remove(previous_year)
                                    _student.academic_years.remove(current_year)
                                    undo_repeated_students.append(_student)
                            old_classes_to_update.append(promoted_students_class)
                        new_classes_to_update.append(_class)
                    except LinkedClasse.DoesNotExist:
                        transaction.set_rollback(True)
                        return Response({'message': f"No class has been linked to the {_class} class"}, status=400)
                    except Exception:
                        transaction.set_rollback(True)
                        return Response(status=400)
                    
                Student.objects.bulk_update(undo_graduated_students, ['current_year', 'has_completed', 'graduation_year', 'graduation_date', 'st_class'])
                Student.objects.bulk_update(undo_promoted_students, ['current_year', 'academic_years', 'st_class', 'linked_classes'])
                Student.objects.bulk_update(undo_repeated_students, ['repeated_academic_years', 'academic_years'])
                Classe.objects.bulk_update(graduated_students_old_classes, ['students', 'name', 'students_year', 'head_teacher', 'program', 'subjects'])
                Classe.objects.bulk_update(old_classes_to_update, ['students'])
                Classe.objects.bulk_update(new_classes_to_update, ['students'])
                for _class in graduated_classes_to_delete:
                    _class.delete()
                
            except Exception:
                transaction.set_rollback(True)
                return Response(status=400)
            
    return Response()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_staff(request):
    data = request.data
    sch_admin = request.user.staff
    school = sch_admin.school
    
    if data['type'] == 'createWithoutFile':
        department = None
        staff_id = None
        first_name = data['firstName']
        last_name = data['lastName']
        role = data['role']
        gender = data['gender']
        img =  data['img'] if data['img'] else None
        contact = data['contact']
        address = data['address']
        date_enrolled = data['dateEnrolled']
        region = data['region']
        religion = data['religion']
        alt_contact = data['altContact']
        pob = data['pob']
        email = data['email']
        nationality = data['nationality']
        subjects = json.loads(data['subjects'])
        dob = data['dateOfBirth']
        if school.has_departments:
            department = Department.objects.get(school=school, name=data['department'])

        user_no = random.randint(100, 999)
        username = f"{data['firstName'].replace(' ', '')[0].upper()}{data['lastName'].replace(' ', '').lower()}{user_no}"
        if school.staff_id:
            staff_id = data['staff_id']
        else:
            staff_id = username
        
        if Staff.objects.filter(school=school, staff_id=staff_id).exists():
            return Response({'message': f"Staff with ID [ {staff_id} ] already exists"}, status=400)
        
        with transaction.atomic():
            try:
                user = User.objects.create_user(
                    username=username,
                    password=username,
                    first_name=first_name,
                    last_name=last_name,
                )
                staff_obj = Staff.objects.create(
                    user=user,
                    staff_id=staff_id,
                    department=department,
                    role=role,
                    gender=gender,
                    dob=dob,
                    school=school,
                    alt_contact=alt_contact,
                    is_active=True,
                    contact=contact,
                    address=address,    
                    email=email,
                    region=region,
                    religion=religion,
                    pob=pob,
                    date_enrolled=date_enrolled,
                    nationality=nationality,
                    img=img,
                )
                subjects_objs = Subject.objects.filter(schools=school, name__in=subjects).distinct()
                staff_obj.subjects.set(subjects_objs)
                user.save()
                staff_obj.save()
                if department:
                    department.teachers.add(staff_obj)
                    department.save()

            except Exception as e:
                print(e)
                transaction.set_rollback(True)
                return Response(status=400)
        
        staff_data = StaffSerializerOne(Staff.objects.get(school=school, staff_id=staff_id)).data
        return Response(staff_data, status=200)

    elif data['type'] == 'setDepartmentHOD':
        staff_hod = Staff.objects.select_related('department').get(school=school, staff_id=data['staffId'])
        department = Department.objects.prefetch_related('teachers').get(school=school, name=data['department'])
        previous_hod_id = None
        if staff_hod not in department.teachers.all():
            transaction.set_rollback(True)
            return Response({'message': f"The staff you selected is not in the {department} department"}, status=400)
        
        with transaction.atomic():
            try:
                if department.hod:
                    previous_hod = department.hod
                    previous_hod.role = 'teacher'
                    previous_hod_id = previous_hod.staff_id
                    previous_hod.save()
                department.hod = staff_hod
                staff_hod.role = 'hod'
                department.save()
                staff_hod.save()
            except:
                transaction.set_rollback(True)
                return Response(status=400) 
            
        return Response({'previous_hod_id': previous_hod_id}, status=200) 
            
    elif data['type'] == 'getFile':
        filepath = format_staff_creation_file(sch_admin)
        filename = 'staff-creation-file.xlsx'
        response = FileResponse(filepath, as_attachment=True, filename=filename)
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response            

    elif data['type'] == 'createWithFile':
        setting_up_start_time = time.time()
        df = None
        title_options = ['Mr', 'Miss', 'Mrs', 'Dr', 'Prof', 'Rev', 'Sir', 'Hon', 'Madam', 'Pastor', 'Imam', 'Bishop', 'Archbishop']
        role_options = ['teacher', 'hod', 'head master', 'head mistress', 'assistant head master', 'assistant head mistress', 'head of academics', 'assistant head of academics', 'administrator']
        religion_options = ['Christianity', 'Islam', 'Traditional African Religions', 'Buddhism', 'Hinduism', 'Sikhism', 'Judaism', 'Other', 'None']
        skipped_rows = 0
        date_field = DateField()
        email_validator = EmailValidator()
        staff_to_create_ids = set()
        existing_staff_ids = []
        users_to_create = []
        staff_to_create = []
        staff_instances = []
        staff_departments = []
        staff_subjects_mappings = defaultdict(list)
        created_staff = []
        try:
            if school.has_departments:
                df = pd.read_excel(data['file'], header=0, sheet_name=0, skiprows=8)
                skipped_rows = 8
            else:
                df = pd.read_excel(data['file'], header=0, sheet_name=0, skiprows=7)
                skipped_rows = 7
                
            df = df.drop(index=0)
            df = df.dropna(how='all')
        except Exception as e:
            return Response({'message': 'Invalid file! Make sure you upload the excel file you generated'}, status=400)
        
        if school.staff_id:
            existing_staff_ids = [x.staff_id for x in Staff.objects.filter(school=school, is_active=True)]
        def show_error_message(label_name:str, column_name:str, row_idx:int):
            return f"The {label_name} cannot be empty! Check the value in the '{column_name}' column in row {row_idx+2+skipped_rows}"
        
        def get_data(row, type:str):
            first_name = row['FIRST NAME'].strip().title() if row['FIRST NAME'] else None
            last_name = row['LAST NAME (MIDDLE NAME + SURNAME)'].strip().title() if row['LAST NAME (MIDDLE NAME + SURNAME)'] else None
            title = row['TITLE'].strip().replace('.', '').title() if row['TITLE'] else None
            staff_id = row['STAFF ID'].strip() if school.staff_id and row.get('STAFF ID') else None
            gender = row['GENDER'].strip().lower() if row['GENDER'] else None
            dob = row['DATE OF BIRTH'].strip() if row['DATE OF BIRTH'] else None
            role = row['ROLE'].strip().lower() if row['ROLE'] else None
            subjects = row['SUBJECT(S)'].strip() if row['SUBJECT(S)'] else None
            department = row['DEPARTMENT'].strip() if school.has_departments and row.get('DEPARTMENT') else None
            date_enrolled = row['DATE EMPLOYED(OPTIONAL)'].strip() if row['DATE EMPLOYED(OPTIONAL)'] else None
            religion = row['RELIGION'].strip().title() if row['RELIGION'] else None
            nationality = row['NATIONALITY'].strip().title() if row['NATIONALITY'] else None
            region = row['REGION/STATE'].strip() if row['REGION/STATE'] else None
            pob = row['PLACE OF BIRTH'].strip() if row['PLACE OF BIRTH'] else None
            contact = row['PHONE NUMBER'].strip()
            alt_contact = row['SECOND PHONE NUMBER(OPTIONAL)'].strip() if row['SECOND PHONE NUMBER(OPTIONAL)'] else None
            address = row['RESIDENTIAL ADDRESS'].strip() if row['RESIDENTIAL ADDRESS'] else None
            email = row['EMAIL(OPTIONAL)'].strip() if row['EMAIL(OPTIONAL)'] else None
            if not first_name:
                error_message = show_error_message(first_name, 'first name', 'FIRST NAME', index)
                raise ErrorMessageException(error_message)
            elif not last_name:
                error_message = show_error_message(last_name, 'last name', 'LAST NAME', index)
                raise ErrorMessageException(error_message)
            elif not gender:
                error_message = show_error_message(gender, 'gender', 'GENDER', index)
                raise ErrorMessageException(error_message)
            elif gender not in ['male', 'female']:
                error_message = f"'{gender}' is not a valid gender! Check the value in the 'GENDER' column in row {index+2+skipped_rows}. Valid values are {['Male', 'Female']}"
                raise ErrorMessageException(error_message)
            elif not dob:
                error_message = show_error_message(dob, 'date of birth', 'DATE OF BIRTH', index)
                raise ErrorMessageException(error_message)
            elif not title:
                error_message = show_error_message(title, 'title', 'TITLE', index)
                raise ErrorMessageException(error_message)
            elif title not in title_options:
                error_message = f"'{title}' is not a valid title! Check the value in the 'TITLE' column in row {index+2+skipped_rows}. Valid values are {title_options}"
                raise ErrorMessageException(error_message)
            elif school.staff_id and not staff_id:
                error_message = show_error_message(title, 'title', 'TITLE', index)
                raise ErrorMessageException(error_message)
            elif not role:
                error_message = show_error_message(role, 'role', 'ROLE', index)
                raise ErrorMessageException(error_message)
            elif role not in role_options:
                error_message = f"'{role}' is not a valid role! Check the value in the 'ROLE' column in row {index+2+skipped_rows}. Valid values are {[x.title() for x in role_options]}"
                raise ErrorMessageException(error_message)
            elif school.has_departments and not department:
                error_message = show_error_message(department, 'department', 'DEPARTMENT', index)
                raise ErrorMessageException(error_message)
            elif not religion:
                error_message = show_error_message(religion, 'religion', 'RELIGION', index)
                raise ErrorMessageException(error_message)
            elif religion not in religion_options:
                error_message = f"'{religion}' is not a valid religion! Check the value in the 'RELIGION' column in row {index+2+skipped_rows}. Valid values are {religion_options}"
                raise ErrorMessageException(error_message)
            elif not nationality:
                error_message = show_error_message(nationality, 'nationality', 'NATIONALITY', index)
                raise ErrorMessageException(error_message)
            elif not validate_nationality(nationality):
                error_message = f"'{nationality}' is not a valid nationality! Check the value in the 'NATIONALITY' column in row {index+2+skipped_rows}."
                raise ErrorMessageException(error_message)
            elif not region:
                error_message = show_error_message(region, 'region', 'REGION/STATE', index)
                raise ErrorMessageException(error_message)
            elif not validate_country_region(get_country_from_nationality(nationality), region):
                error_message = f"'{region}' is not a valid region/state in {nationality}! Check the value in the 'REGION/STATE' column in row {index+2+skipped_rows}."
                raise ErrorMessageException(error_message)
            elif not pob:
                error_message = show_error_message(pob, 'place of birth', 'PLACE OF BIRTH', index)
                raise ErrorMessageException(error_message)
            elif not contact :
                error_message = show_error_message(contact, 'phone number', 'PHONE NUMBER', index)
                raise ErrorMessageException(error_message)
            elif not valid_phone_number(contact):
                error_message = f"'{contact}' is not a valid phone number! Check the value in the 'PHONE NUMBER' column in row {index+2+skipped_rows}."
                raise ErrorMessageException(error_message)
            elif alt_contact and not valid_phone_number(alt_contact):
                error_message = f"'{alt_contact}' is not a valid phone number! Check the value in the 'SECOND PHONE NUMBER' column in row {index+2+skipped_rows}."
                raise ErrorMessageException(error_message)
            elif not address :
                error_message = show_error_message(address, 'residential addresss', 'RESIDENTIAL ADDRESS', index)
                raise ErrorMessageException(error_message)
            
            try:
                dob = datetime.strptime(dob, "%d-%m-%Y").strftime("%Y-%m-%d")
                date_field.clean(dob)
            except Exception:
                error_message = f"'{dob}' is not a valid date! Check the value in the 'DATE OF BIRTH' column in row {index+2+skipped_rows}."
                raise ErrorMessageException(error_message)
            
            if school.staff_id:
                if staff_id in staff_to_create_ids:
                    error_message = f"Two staff cannot have the same staff ID. The staff ID in row {index+2+skipped_rows} conflicts with another staff ID on the sheet"
                    raise ErrorMessageException(error_message)
                elif staff_id in existing_staff_ids:
                    error_message = f"Staff with ID '{staff_id}' already exists. Check the staff ID in row {index+2+skipped_rows}"
                    raise ErrorMessageException(error_message)
            else:
                staff_id = username
                    
            if date_enrolled:
                try:
                    date_enrolled = datetime.strptime(date_enrolled, "%d-%m-%Y").strftime("%Y-%m-%d")
                    date_field.clean(date_enrolled)
                except Exception:
                    error_message = f"'{date_enrolled}' is not a valid date! Check the value in the 'DATE EMPLOYED(OPTIONAL)' column in row {index+2+skipped_rows}."
                    raise ErrorMessageException(error_message)
                    
            if email:
                try:
                    email_validator(email)
                except Exception:
                    error_message = f"'{email}' is not a valid email address! Check the value in the 'EMAIL(OPTIONAL)' column in row {index+2+skipped_rows}."
                    raise ErrorMessageException(error_message)
            
            if subjects:
                try:
                    subjects_objs = Subject.objects.filter(schools=school, name__in=[x.strip().upper() for x in subjects.split(',')]).distinct()
                    if len(subjects_objs) == 0:
                        raise(ValueError)
                    else:
                        subjects = subjects_objs
                except Exception as e:
                    error_message = f"'{subjects}' is/are not a valid subject(s)! Check the value in the 'SUBJECT(S)' column in row {index+2+skipped_rows}."
                    raise ErrorMessageException(error_message)
                    
            if school.has_departments:
                try:
                    department = Department.objects.get(school=school, name=department.strip().upper())
                except Exception as e:
                    error_message = f"'{department}' is not a valid department! Check the value in the 'DEPARTMENT' column in row {index+2+skipped_rows}."
                    raise ErrorMessageException(error_message)
            
            user_no = random.randint(100, 999)
            username = f"{first_name.replace(' ', '')[0].upper()}{last_name.replace(' ', '').lower()}{user_no}"
            if type == 'user':
                return {
                    'first_name': first_name,
                    'last_name': last_name,
                    'username': username,
                }
            elif type == 'staff':
                return {
                    'staff_id': staff_id,
                    'title': f"{title.title()}.",
                    'department': department,
                    'role': role,
                    'gender': gender,
                    'subjects': subjects,
                    'dob': dob,
                    'date_enrolled': date_enrolled,
                    'religion': religion,
                    'region': region,
                    'nationality': nationality,
                    'pob': pob,
                    'contact': contact,
                    'alt_contact': alt_contact,
                    'email': email,
                    'address': address,
                }
            else:
                raise Exception("The type must be either 'user' or 'staff'")

        setting_up_end_time = time.time()
        print(f"Setting up time takend: {setting_up_end_time - setting_up_start_time}")
        
        with transaction.atomic():
            try:
                user_start_time = time.time()
                for index, row in df.iterrows():
                    user_data = get_data(row, 'user')
                    staff_data = get_data(row, 'staff')
                    user = User(
                        username=user_data['username'],
                        first_name=user_data['first_name'],
                        last_name=user_data['last_name'],
                    )
                    user.set_password(user_data['username'])
                    users_to_create.append(user)
                    staff_to_create.append((user_data['username'], staff_data))
                    staff_to_create_ids.add(staff_data['staff_id'])
                    
                User.objects.bulk_create(users_to_create)
                created_users = User.objects.filter(username__in=[user.username for user in users_to_create])
                user_mapping = {user.username: user for user in created_users}
                user_end_time = time.time()
                print(f"User creation time: {user_end_time - user_start_time}")
                staff_start_time = time.time()
                for username, staff_data in staff_to_create:
                    user = user_mapping[username]
                    staff = Staff(
                        user=user,
                        school=school,
                        staff_id=staff_data['staff_id'],
                        title=staff_data['title'],
                        department=staff_data['department'],
                        role=staff_data['role'],
                        gender=staff_data['gender'],
                        dob=staff_data['dob'],
                        date_enrolled=staff_data['date_enrolled'],
                        religion=staff_data['religion'],
                        region=staff_data['region'],
                        nationality=staff_data['nationality'],
                        pob=staff_data['pob'],
                        contact=staff_data['contact'],
                        alt_contact=staff_data['alt_contact'],
                        email=staff_data['email'],
                        address=staff_data['address'],
                        is_active=True,
                    )
                    staff_subjects_mappings[staff_data['staff_id']] = staff_data['subjects']
                    staff_instances.append(staff)
                    if school.has_departments and staff_data['department']:
                        staff_departments.append(staff_data['department'])
        
                Staff.objects.bulk_create(staff_instances)
                created_staff = Staff.objects.select_related('department').filter(school=school, staff_id__in=staff_to_create_ids)
                staff_end_time = time.time()
                print(f"Staff creation time: {staff_end_time - staff_start_time}")
                
                for _staff in created_staff:
                    staff_subjects = staff_subjects_mappings[_staff.staff_id]
                    if staff_subjects:
                        _staff.subjects.set(staff_subjects)
                    if _staff.department and _staff.department in staff_departments:
                        department_index = staff_departments.index(_staff.department)
                        staff_departments[department_index].teachers.add(_staff)
            
            except ErrorMessageException as e:
                return Response({'message': str(e)}, status=400)
            except Exception as e:
                    transaction.set_rollback(True)
                    print(e)
                    return Response(status=400)
        
        staff_data = StaffSerializerOne(created_staff, many=True).data
        return Response(staff_data, status=200)

    elif data['type'] == 'delete':
        staff_to_delete = Staff.objects.get(school=school, staff_id=data['staffId'])
        user = staff_to_delete.user
        if SubjectAssignment.objects.filter(school=school, teacher=staff_to_delete).exists():
            return Response({'message': "You don't have permission to delete this staff"}, status=400)
        
        students_classes = Classe.objects.filter(school=school)
        for _class in students_classes:
            if _class.head_teacher == staff_to_delete:
                return Response({'message': "You don't have permission to delete this staff"}, status=400)
        
        with transaction.atomic():
            try:
                user.delete()
                staff_to_delete.delete()
            except Exception as e:
                transaction.set_rollback(True)
                return Response(status=400)

        return Response(status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_students(request):
    sch_admin = request.user.staff
    school = sch_admin.school
    data = request.data
    
    if data['type'] == 'createWithoutFile':
        first_name = data['firstName']
        last_name = data['lastName']
        dob = data['dob']
        contact = data['contact']
        address = data['address']
        guardian_contact = data['guardianContact']
        guardian_address = data['guardianAddress']
        guardian_email = data['guardianEmail']
        guardian_nationality = data['guardianNationality']
        guardian_occupation = data['guardianOccupation']
        date_enrolled = data['dateEnrolled']
        img = data['img']
        email = data['email']
        gender = data['gender']
        nationality = data['nationality']
        guardian_fullname = data['guardianFullname']
        guardian_gender = data['guardianGender']
        religion = data['religion']
        region = data['region']
        pob = data['pob']
        student_id = None
        user_no = random.randint(100, 999)
        username = f"{first_name.replace(' ', '')[0].upper()}{last_name.replace(' ', '').lower()}{user_no}"
        
        if school.students_id:
            student_id = data['studentId']
        else:
            student_id = username

        try:
            Student.objects.get(school=school, st_id=student_id)
            return Response({'message': f'Student with ID {student_id} already exists'}, status=400)
        except Student.DoesNotExist:
            pass
        
        years_to_complete = school.level.years_to_complete
        academic_years = AcademicYear.objects.filter(school=school).order_by('-start_date')
        current_year = academic_years[0]
        student_class = Classe.objects.select_related('program').get(school=school, name=data['className'])
        with transaction.atomic():
            try:
                user = User.objects.create_user(
                    username=username,
                    password=username,
                    first_name=first_name,
                    last_name=last_name,
                )
                student_obj = Student.objects.create(
                    school=school,
                    user=user,
                    program=student_class.program,
                    st_class=student_class,
                    current_year=student_class.students_year,
                    st_id=student_id,
                    gender=gender,
                    date_enrolled=date_enrolled,
                    dob=dob,
                    img=img if img else None,
                    contact=contact,
                    email=email,
                    address=address,
                    nationality=nationality,
                    pob=pob,
                    graduation_date=current_year.students_graduation_date if student_class.students_year == years_to_complete else None,
                    religion=religion,
                    region=region,
                    guardian=guardian_fullname,
                    guardian_address=guardian_address,
                    guardian_gender=guardian_gender,
                    guardian_contact=guardian_contact,
                    guardian_email=guardian_email,
                    guardian_occupation=guardian_occupation,
                    guardian_nationality=guardian_nationality,
                )
                academic_year_count = 0
                for _year in range(student_class.students_year):
                    student_obj.academic_years.add(academic_years[academic_year_count])
                    academic_year_count += 1
                
                student_class.students.add(student_obj)
                user.save()
                student_obj.save()
                student_class.save()
           
            except Exception as e:
                transaction.set_rollback(True)
                return Response(status=400)
        
        student_data = StudentSerializerOne(Student.objects.get(school=school, st_id=student_id)).data
        return Response(student_data, status=200)

    elif data['type'] == 'get-students-file':
        sch_admin_data = StaffSerializer(sch_admin).data
        class_name = data['className'].split()[0]
        clas = Classe.objects.filter(school=sch_admin.school, name=class_name).first()
        if clas:
            clas_data = ClasseSerializer(clas).data
            wb = load_workbook('staticfiles/files/students_admission.xlsx')
            ws = wb.worksheets[0]
            filename = f"{clas_data['name']}-FORM-{clas_data['students_year']}.xlsx"
            ws.title = f"{clas_data['name']} FORM-{clas_data['students_year']}"
            byte_file = io.BytesIO()
            wb.save(byte_file)

            if settings.DEBUG:
                file_path = f"{get_school_folder(sch_admin_data['school']['name'])}/staff/{sch_admin_data['user']['username']}/{filename}"
                if default_storage.exists(file_path):
                    default_storage.delete(file_path)

                save_file = default_storage.save(file_path, byte_file)

                return Response({
                    'filename': f"{filename}.xlsx",
                    'file_path': f"http://localhost:8000{default_storage.url(save_file)}",
                })

            else:
                file_path = f"{get_school_folder(sch_admin_data['school']['name'])}/staff/{sch_admin_data['user']['username']}/{filename}"
                if default_storage.exists(file_path):
                    default_storage.delete(file_path)

                save_file = default_storage.save(file_path, byte_file)

                return Response({
                    'filename': f"{filename}.xlsx",
                    'file_path': default_storage.url(save_file),
                })

    elif data['type'] == 'upload-students-file':
        class_name = data['className'].split()[0]
        wb = load_workbook(data['file'])
        if data['className'] == wb.sheetnames[0]:
            wb.close()
            df = pd.read_excel(data['file'], sheet_name=data['className'])
            df_data = df.iloc[2:, :5].dropna().values.tolist()

            existing_student_ids = set(Student.objects.filter(
                school=sch_admin.school,
            ).values_list('st_id', flat=True))

            with transaction.atomic():
                clas = Classe.objects.select_related('program').get(school=sch_admin.school, name=class_name)
                user_no = random.randint(100, 999)

                for st in df_data:
                    if str(st[2]) in existing_student_ids:
                        transaction.set_rollback(True)
                        return Response({'ms': f"Student with ID {st[2]} already exists"}, status=201)
                    
                    username = f"{st[0].replace(' ', '')[0].upper()}{st[1].replace(' ', '').lower()}{user_no}"
                    user = User.objects.create_user(
                        username=username,
                        password=username,
                        first_name=st[0].upper(),
                        last_name=st[1].upper(),
                    )
                    user.save()

                    st_user = User.objects.get(username=username)

                    try:
                        student = Student.objects.create(
                            user=st_user,
                            program=clas.program,
                            st_id=st[2],
                            current_year=clas.students_year,
                            st_class=clas,
                            date_enrolled=clas.date_enrolled,
                            gender=st[3].upper(),
                            dob=st[4],
                            school=sch_admin.school,
                        )
                        student.save()
                        clas.students.add(student)

                    except ValidationError:
                        transaction.set_rollback(True)
                        return Response({
                            'ms': f"{st[4]} is not a valid date, recheck the date of birth of {st[0]} {st[1]} in the file"},
                            status=201)
                        
                    except IntegrityError:
                        transaction.set_rollback(True)
                        return Response({
                            'ms': f"Student ID {st[2]} has been assigned to two students, recheck the students ids in the file"},
                            status=201)

                clas.save()

            class_data = ClasseSerializer(Classe.objects.get(school=sch_admin.school, name=class_name)).data
            return Response({
                'ms': 'File uploaded and students saved successfully',
                'data': class_data,
                'year': 'yearOne' if class_data['students_year'] == 1 else (
                    'yearTwo' if class_data['students_year'] == 2 else 'yearThree'),
            })

        else:
            return Response({'ms': 'The first sheet name must be same as the selected class'}, status=201)

    elif data['type'] == 'delete-student':
        st_data = SpecificStudentSerializer(Student.objects.get(st_id=data['stId'])).data
        class_data = ClasseSerializer(Classe.objects.get(name=data['className'])).data
        user = User.objects.get(username=st_data['user']['username'])

        with transaction.atomic():
            # delete user object
            user.delete()
            student = Student.objects.get(st_id=data['stId'])
            clas = Classe.objects.get(name=data['className'])
            # remove student from class
            clas.students.remove(student)
            clas.save()
            # delete student object
            student.delete()

        return Response({
            'class_name': class_data['name'],
            'year': 'yearOne' if class_data['students_year'] == 1 else (
                'yearTwo' if class_data['students_year'] == 2 else 'yearThree'),
        })

    elif data['type'] == 'delete-subject':
        subject = Subject.objects.get(name=data['subject'])
        clas = Classe.objects.get(school=sch_admin.school, name=data['className'])
        class_data = ClasseWithoutStudentsSerializer(clas).data

        with transaction.atomic():
            # remove subject from class
            clas.subjects.remove(subject)
            clas.save()

        return Response({
            'class_name': class_data['name'],
            'year': 'yearOne' if class_data['students_year'] == 1 else (
                'yearTwo' if class_data['students_year'] == 2 else 'yearThree'),
        })

    elif data['type'] == 'delete-class':
        clas = Classe.objects.get(school=sch_admin.school, name=data['className'])

        with transaction.atomic():
            # remove subject from class
            clas.delete()

        return Response(status=200)
    
    

