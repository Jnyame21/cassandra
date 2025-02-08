# Django
from django.db import IntegrityError, transaction
from django.core.files.storage import default_storage

# Document Manipulation
import pandas as pd
from openpyxl.styles import Font, Alignment, Protection
from openpyxl import Workbook, load_workbook

# Django Restframework
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from api.models import *
from api.serializer import *
from api.utils import *
import json
from collections import defaultdict
from datetime import datetime
import hashlib
import inflect
import math
from django.db.models import Prefetch
from decimal import Decimal
import time

# TEACHER 
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_teacher_data(request):
    teacher = Staff.objects.select_related('school', 'current_role__level').get(user=request.user)
    school = teacher.school
    current_level = teacher.current_role.level
    current_term = int(request.GET.get('term'))
    current_academic_year = AcademicYear.objects.get(id=int(request.GET.get('year')))
    staff = StaffSerializerTwo(Staff.objects.select_related('user').prefetch_related('departments', 'subjects', 'roles').filter(school=school), many=True).data
    subject_assignments_objs = SubjectAssignment.objects.prefetch_related('subjects', 'students_class__students__user').filter(school=school, level=current_level, teacher=teacher, academic_year=current_academic_year, academic_term=current_term)
    subject_assignments_data = {}
    all_students_assessments_data = {}
    all_students_exams_data = {}
    all_students_results_data = {}
    all_students_assessments_objs = Assessment.objects.select_related('student__user', 'student_class', 'subject').filter(school=school, level=current_level, teacher=teacher, academic_year=current_academic_year, academic_term=current_term).order_by('-score')
    assessment_by_class_subject = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for _assessment in all_students_assessments_objs:
        assessment_by_class_subject[_assessment.student_class.name][_assessment.subject.name][_assessment.title].append(_assessment)
    
    all_students_exams_objs = Exam.objects.select_related('student__user', 'student_class', 'subject').filter(school=school, level=current_level, teacher=teacher, academic_year=current_academic_year, academic_term=current_term).order_by('-score')
    exams_by_class_subject = defaultdict(lambda: defaultdict(list))
    for _exam in all_students_exams_objs:
        exams_by_class_subject[_exam.student_class.name][_exam.subject.name].append(_exam)
    
    all_students_results_objs = StudentResult.objects.select_related('student__user', 'student_class', 'subject').filter(school=school, level=current_level, teacher=teacher, academic_year=current_academic_year, academic_term=current_term).order_by('-result')
    results_by_class_subject = defaultdict(lambda: defaultdict(list))
    for _result in all_students_results_objs:
        results_by_class_subject[_result.student_class.name][_result.subject.name].append(_result)
    
    for assign in subject_assignments_objs:
        students_class = assign.students_class
        students_class_name = students_class.name
        class_students = students_class.students.all()
        subject_assignments_data[students_class_name] = SubjectAssignmentSerializerTwo(assign).data
        
        assessment_class_data = {}
        exams_class_data = {}
        results_class_data = {}
        assignment_subjects = assign.subjects.all()
        for _subject in assignment_subjects:
            
            ## Students Assessments
            assessment_subject_data = {}
            assessment_titles = set(assessment_by_class_subject[students_class_name][_subject.name].keys())
            for _title in assessment_titles:
                assessments_with_title = assessment_by_class_subject[students_class_name][_subject.name][_title]
                first_assessment = assessments_with_title[0]
                assessment_title_data = {
                    'title': _title, 
                    'total_score': first_assessment.total_score, 
                    'description': first_assessment.description, 
                    'percentage': first_assessment.percentage, 
                    'assessment_date': first_assessment.assessment_date, 
                    'students_with_assessment': {}, 
                    'students_without_assessment': {}
                }
                students_with_assessments = set()
                for _assessment in assessments_with_title:
                    assessment_student = _assessment.student
                    if assessment_student:
                        students_with_assessments.add(assessment_student.st_id)
                        student_data = {
                            'name': f"{assessment_student.user.first_name} {assessment_student.user.last_name}",
                            'st_id': assessment_student.st_id,
                            'score': _assessment.score,
                            'comment': _assessment.comment,
                        }
                        assessment_title_data['students_with_assessment'][assessment_student.st_id] = student_data
                        
                for _student in class_students:
                    st_id = _student.st_id
                    if st_id not in students_with_assessments:
                        st_data = {
                            'name': f"{_student.user.first_name} {_student.user.last_name}",
                            'st_id': st_id,
                        }
                        assessment_title_data['students_without_assessment'][st_id] = st_data
                assessment_subject_data[_title] = assessment_title_data
            assessment_class_data[_subject.name] = assessment_subject_data
            
            # Students Exams
            exams_subject_data = {'students_with_exams': {}, 'students_without_exams': {}}
            all_subject_exams_objs = exams_by_class_subject[students_class_name][_subject.name]
            students_with_exams = set()
            if len(all_subject_exams_objs) > 0:
                first_exams_objs = all_subject_exams_objs[0]
                exams_subject_data['percentage'] = first_exams_objs.percentage
                exams_subject_data['total_score'] = first_exams_objs.total_score
                for _exams in all_subject_exams_objs:
                    exams_student = _exams.student
                    if exams_student:
                        student_data = {
                            'name': f"{exams_student.user.first_name} {exams_student.user.last_name}",
                            'st_id': exams_student.st_id,
                            'score': _exams.score,
                        }
                        exams_subject_data['students_with_exams'][exams_student.st_id] = student_data
                        students_with_exams.add(exams_student.st_id)
            
                for _student in class_students:
                    st_id = _student.st_id
                    if st_id not in students_with_exams:
                        student_data = {
                            'name': f"{_student.user.first_name} {_student.user.last_name}",
                            'st_id': st_id,
                        }
                        exams_subject_data['students_without_exams'][st_id] = student_data
            else:
                exams_subject_data['percentage'] = 0
                exams_subject_data['total_score'] = 0
                
            exams_class_data[_subject.name] = exams_subject_data
            
            # Students Results
            results_subject_data = {}
            all_subject_results_objs = results_by_class_subject[students_class_name][_subject.name]
            if len(all_subject_results_objs) > 0:
                first_result_obj = all_subject_results_objs[0]
                results_subject_data['total_assessment_percentage'] = first_result_obj.total_assessment_percentage
                results_subject_data['exam_percentage'] = first_result_obj.exam_percentage
                all_results = {}
                for _result in all_subject_results_objs:
                    result_student = _result.student
                    student_data = {
                        'result': _result.result,
                        'total_assessment_score': _result.total_assessment_score,
                        'exam_score': _result.exam_score,
                        'student': {'name': f"{_result.student.user.first_name} {_result.student.user.last_name}", 'st_id': result_student.st_id},
                        'remark': _result.remark,
                        'grade': _result.grade,
                        'position': _result.position,
                    }
                    all_results[result_student.st_id] = student_data
                results_subject_data['student_results'] = all_results
            
            else:
                results_subject_data['total_assessment_percentage'] = 0
                results_subject_data['exam_percentage'] = 0
                results_subject_data['student_results'] = {}
                
            results_class_data[_subject.name] = results_subject_data
            
        # All(assessments, exams, results) data for the class
        all_students_assessments_data[students_class_name] = assessment_class_data
        all_students_exams_data[students_class_name] = exams_class_data
        all_students_results_data[students_class_name] = results_class_data
            
    
    students_class_objs = Classe.objects.select_related('head_teacher').prefetch_related('students__user').filter(school=school, level=current_level).order_by('-students_year')
    students_class_names = []
    head_classes = []
    for _class in students_class_objs:
        if _class.head_teacher == teacher:
            head_classes.append(_class)
        students_class_names.append(_class.name)
        
    attendance_mappings = defaultdict(list)
    attendance_objs = StudentAttendance.objects.select_related('students_class').prefetch_related('students_present__user', 'students_absent__user').filter(school=school, level=current_level, students_class__in=head_classes, academic_year=current_academic_year, academic_term=current_term).order_by('-date')
    for _attendance_ in attendance_objs:
        attendance_mappings[_attendance_.students_class.id].append(_attendance_)
    atttendance_data = defaultdict(dict)
    for _class in head_classes:
        students_class_name = _class.name
        students = StudentUserIdSerializer(_class.students.all(), many=True).data
        attendances = StudentsAttendanceSerializer(attendance_mappings[_class.id], many=True).data
        atttendance_data[students_class_name]['students'] = students
        atttendance_data[students_class_name]['attendances'] = attendances
    
    department_data = None
    hod_subject_assignments_data = []
    department = Department.objects.select_related('hod__user').prefetch_related('subjects', 'teachers__user').filter(school=school, level=current_level, teachers=teacher).first()
    if department:
        department_data = DepartmentSerializerOne(department).data
        if department.hod == teacher:
            hod_subject_assignments = SubjectAssignment.objects.select_related('students_class', 'teacher__user').prefetch_related('subjects').filter(school=school, level=current_level, assigned_by=teacher, academic_year=current_academic_year, academic_term=current_term)
            hod_subject_assignments_data = SubjectAssignmentSerializerOne(hod_subject_assignments, many=True).data
    
    return Response({
        'subject_assignments': subject_assignments_data,
        'department_data': department_data,
        'staff': staff,
        'students_attendance': atttendance_data,
        'students_assessments': all_students_assessments_data,
        'students_exams': all_students_exams_data,
        'students_results': all_students_results_data,
        'hod_subject_assignments': hod_subject_assignments_data,
        'hod_student_classes': students_class_names,
    })
    
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def teacher_students_attendance(request):
    teacher = Staff.objects.select_related('school', 'current_role__level').get(user=request.user)
    school = teacher.school
    current_level = teacher.current_role.level
    data = request.data
    students_class = Classe.objects.prefetch_related('students').get(school=school, level=current_level, name=data['className'])
    current_academic_year = AcademicYear.objects.get(school=school, level=current_level, name=data['year'])
    current_term = int(data['term'])
    
    if data['type'] == 'create':
        student_ids = data['absentStudents'].split(',')
        student_objs = students_class.students.all()
        if StudentAttendance.objects.filter(school=school, level=current_level, date=data['date'], teacher=teacher, students_class=students_class, academic_year=current_academic_year, academic_term=current_term).exists():
            return Response({'message': f"You have already uploaded attendance for this date [ {data['date']} ]"}, status=400)

        with transaction.atomic():
            attendance = StudentAttendance.objects.create(
                school=school,
                level=current_level,
                teacher=teacher,
                date=data['date'],
                students_class=students_class,
                academic_year=current_academic_year,
                academic_term=int(data['term']),
            )
                
            for st in student_objs:
                if st.st_id not in student_ids:
                    attendance.students_present.add(st)
                else:
                    attendance.students_absent.add(st)

        attendance_obj = StudentAttendance.objects.prefetch_related('students_absent__user', 'students_present__user').get(school=school, level=current_level, academic_year=current_academic_year, academic_term=current_term, date=data['date'], students_class=students_class, teacher=teacher)
        attendance_data = StudentsAttendanceSerializer(attendance_obj).data
        return Response(attendance_data, status=200)

    elif data['type'] == 'delete':
        try:
            attendance_id = int(data['id'])
            attendance = StudentAttendance.objects.select_related('academic_year').get(id=attendance_id)
            if attendance.academic_year != current_academic_year and attendance.academic_term != current_term:
                return Response({'message': f"You don't have permission to delete this attendance"}, status=400)
            
            with transaction.atomic():
                attendance.delete()

            return Response(status=200)

        except StudentAttendance.DoesNotExist:
            return Response({'message': 'Attendance data not found'}, status=400)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def teacher_assessments(request):
    teacher = Staff.objects.select_related('school', 'current_role__level').get(user=request.user)
    school = teacher.school
    current_level = teacher.current_role.level
    data = request.data
    subject_name = data['subject']
    st_class_name = data['studentsClassName']
    year = data['year']
    current_academic_year = AcademicYear.objects.get(school=school, level=current_level, name=year)
    term = int(data['term'])
    assessment_title = data['title']
    
    if data['type'] == 'getFile':
        students = json.loads(data['selectedStudents'])
        total_score = data['totalScore']
        if current_level.students_id:
            data = [
                ['STUDENT NAME', 'STUDENT ID', 'SCORE', 'COMMENT']
            ]
        else:
            data = [
                ['STUDENT NAME', 'STUDENT USERNAME', 'SCORE', 'COMMENT']
            ]
            
        if len(students) == 0:
            return Response({'message': f"You have already uploaded all the {subject_name}[{assessment_title}] Assessment scores for all students in this class"}, status=400)
            
        for _st in students:
            row = [_st['name'], _st['st_id'], '']
            data.append(row)

        wb = Workbook()
        ws = wb.active
        ws.title = subject_name

        ws.merge_cells('A1:I3')
        ws['A1'].value = f"SUBJECT: [{subject_name}]  ACADEMIC YEAR: [{year}]  CLASS: [{st_class_name}]  {current_academic_year.period_division}: [{term}]  ASSESSMENT: [{assessment_title}]"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        count = 0
        container = ws['B5:E200']
        while count < len(data):
            for row_ws, row_data in zip(container, data):
                for cell, item in zip(row_ws, row_data):
                    cell.value = item

            count += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50

        for row in ws['A3:F200']:
            for cell in row:
                cell.font = Font(size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws['B3:B200']:
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for cells in ws['A5:F200'][0]:
            cells.font = Font(bold=True, size=14)

        ws.protection.sheet = True
        for cell in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=4, max_col=4):
            cell[0].protection = Protection(locked=False)
        for cell in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=5, max_col=5):
            cell[0].protection = Protection(locked=False)

        ws.protection.password = 'teamjn'

        filename = f"{subject_name.replace(' ', '-')}-{st_class_name}-{assessment_title}"
        byte_file = io.BytesIO()
        wb.save(byte_file)
        
        return send_file('excel', byte_file, filename)

    elif data['type'] == 'createAssessment':
        students_class = Classe.objects.get(school=school, level=current_level, name=st_class_name)
        subject_obj = Subject.objects.get(schools=school, level=current_level, name=subject_name)
        description = data['description']
        total_score = float(data['totalScore'])
        assessment_date = data['date']
        assessment_date_object = datetime.strptime(assessment_date, '%Y-%m-%d').date()
        existing_results = StudentResult.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).exists()
        if existing_results:
            return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
        if total_score <= 0:
            return Response({'message': f"The total score of the assessment cannot be negative or zero(0)"}, status=400)
        if (assessment_date_object > current_academic_year.end_date) or (assessment_date_object < current_academic_year.start_date):
            return Response({'message': f"The assessment date must be between the academic year start and end dates. That's {current_academic_year.start_date} and {current_academic_year.end_date}"}, status=400)
        if Assessment.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term).exists():
            return Response({'message': f"An assessment with the title '{assessment_title}' already exists. Please use a different title."}, status=400)
        
        assessment_id = None
        with transaction.atomic():
            try:
                assessment_obj = Assessment.objects.create(
                    school=school, 
                    level=current_level, 
                    teacher=teacher, 
                    subject=subject_obj, 
                    percentage=0,
                    student_class=students_class,
                    title=assessment_title,
                    description=description,
                    total_score=float(total_score),
                    assessment_date=assessment_date,
                    academic_year=current_academic_year, 
                    academic_term=term
                )
                assessment_id = assessment_obj.id
            except Exception:
                return Response(status=400) 
        
        return Response({'message': "Assessment created successfully", 'id': assessment_id})
        
    elif data['type'] == 'uploadWithFile':
        file = data['file']
        subject_obj = Subject.objects.get(name=subject_name)
        description = data['description']
        total_score = float(data['totalScore'])
        percentage = float(data['percentage'])
        assessment_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
            
        try:
            wb = load_workbook(file)
        except Exception:
            return Response({'message': "Invalid file. Ensure you upload the excel file that you generated"}, status=400)
        
        ws = wb.active
        ws.protection = False
        if ws['A1'].value != f"SUBJECT: [{subject_name}]  ACADEMIC YEAR: [{year}]  CLASS: [{st_class_name}]  {current_academic_year.period_division}: [{term}]  ASSESSMENT: [{assessment_title}]":
            return Response({'message': "Invalid file. Ensure you upload the same excel file that you got"}, status=400)
        
        cleaned_data = []
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=5):
            data = []
            for cell in row:
                data.append(cell.value)
            cleaned_data.append(data)
        
        if len(cleaned_data) == 0:
            return Response({'message': f"There are no students data in the uploaded file"}, status=400)

        valid_rows = [x for x in cleaned_data if x[0] and x[1]]
        for row in valid_rows:
            score = row[2]
            try:
                float(score)
            except Exception:
                return Response({'message': f"The scores must be a number. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)
            
            if score > total_score:
                return Response({'message': f"Scores cannot exceed the total score for the assessment. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)
            elif score < 0:
                return Response({'message': f"Scores cannot be negative. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)

        students_class = Classe.objects.prefetch_related('students').get(school=school, level=current_level, name=st_class_name)
        class_students_ids = [x.st_id for x in students_class.students.all()]
        student_ids = [x[1] for x in valid_rows]
        assessment_students_objs = {x.st_id: x for x in students_class.students.filter(st_id__in=student_ids)}
        with transaction.atomic():
            assessments_to_create = []
            for _student in valid_rows:
                st_id = _student[1]
                if st_id in class_students_ids:
                    student = assessment_students_objs[st_id]
                    st_assessment = Assessment(
                        school=school,
                        student=student,
                        subject=subject_obj,
                        teacher=teacher,
                        student_class=students_class,
                        score=float(_student[2]),
                        academic_year=current_academic_year,
                        title=assessment_title,
                        description=description,
                        percentage=percentage,
                        total_score=total_score,
                        comment=_student[3] if _student[3] else '',
                        academic_term=term,
                        level=current_level,
                        assessment_date=assessment_date,
                    )
                    assessments_to_create.append(st_assessment)
                else:
                    transaction.set_rollback(True)
                    return Response({'message': f"Invalid student {_student[0]}[{_student[1]}]. Make sure you don't change anything in the excel file except the scores and comments"}, status=400)
            try:
                Assessment.objects.bulk_create(assessments_to_create)
                assessment_to_delete = Assessment.objects.filter(school=school, level=current_level, teacher=teacher, student=None, student_class=students_class, subject=subject_obj, title=assessment_title, academic_year=current_academic_year, academic_term=term).first()
                if assessment_to_delete:
                    assessment_to_delete.delete()
            except IntegrityError:
                transaction.set_rollback(True)
                return Response({'message': f"You have already uploaded Assessment for some of the students in the file. Click on get file to get a new updated file"}, status=400)
            except Exception:
                transaction.set_rollback(True)
                return Response({'message': f"An unexpected error occurred! Ensure you don't delete or change anything in the excel file except the scores and comments"}, status=400)
        
        assessments = Assessment.objects.select_related('student__user').filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, title=assessment_title, student__in=list(assessment_students_objs.values()), academic_year=current_academic_year, academic_term=term)
        students_data = [{
            'name': f"{_assessment.student.user.first_name} {_assessment.student.user.last_name}", 
            'st_id': _assessment.student.st_id,
            'score': _assessment.score,
            'comment': _assessment.comment,
        } for _assessment in assessments]
        return Response({'message': "Assessment data uploaded successfully!", 'data': students_data})

    elif data['type'] == 'uploadWithoutFile':
        student_ids = json.loads(data['selectedStudents'])
        students_class = Classe.objects.prefetch_related(Prefetch('students', queryset=Student.objects.filter(st_id__in=student_ids))).get(school=school, level=current_level, name=st_class_name)
        subject_obj = Subject.objects.get(name=subject_name)
        comment = data['comment']
        description = data['description']
        total_score = float(data['totalScore'])
        percentage = float(data['percentage'])
        assessment_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        
        score = float(data['score'])
        if score < 0:
            return Response({'message': f"The student(s) score cannot be negative!"}, status=400)
        if score > total_score:
            return Response({'message': f"The student(s) score cannot exceed the total score for the assessment!"}, status=400)
        
        students_objs = {x.st_id: x for x in students_class.students.all()}
        assessments_to_create = []
        with transaction.atomic():
            for _st_id in student_ids:
                student = students_objs[_st_id]
                st_assessment = Assessment(
                    school=school,
                    student=student,
                    subject=subject_obj,
                    teacher=teacher,
                    student_class=students_class,
                    score=score,
                    academic_year=current_academic_year,
                    title=assessment_title,
                    description=description,
                    percentage=percentage,
                    total_score=total_score,
                    comment=comment,
                    academic_term=term,
                    level=current_level,
                    assessment_date=assessment_date,
                )
                assessments_to_create.append(st_assessment)
            
            try:
                Assessment.objects.bulk_create(assessments_to_create)
                assessment_to_delete = Assessment.objects.filter(school=school, level=current_level, teacher=teacher, student=None, student_class=students_class, subject=subject_obj, title=assessment_title, academic_year=current_academic_year, academic_term=term).first()
                if assessment_to_delete:
                    assessment_to_delete.delete()
            except Exception:
                transaction.set_rollback(True)
                return Response(status=400)
        
        assessments = Assessment.objects.select_related('student__user').filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, title=assessment_title, student__in=list(students_objs.values()), academic_year=current_academic_year, academic_term=term)
        students_data = [{
            'name': f"{_assessment.student.user.first_name} {_assessment.student.user.last_name}", 
            'st_id': _assessment.student.st_id,
            'score': _assessment.score,
            'comment': _assessment.comment,
        } for _assessment in assessments]
        return Response({'message': "Assessment data uploaded successfully!", 'data': students_data})

    elif data['type'] == 'editAssessment':
        students_class = Classe.objects.get(school=school, level=current_level, name=st_class_name)
        subject_obj = Subject.objects.get(name=subject_name)
        
        if data['editType'] == 'title':
            with transaction.atomic():
                assessments_to_update = []
                old_title = assessment_title
                new_title = data['newTitle']
                if Assessment.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, title=new_title, academic_year=current_academic_year, academic_term=term).exists():
                    return Response({'message': f"Assessment with title [ {new_title} ] already exists! Use a different title"}, status=400)
                assessments = Assessment.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, title=old_title, academic_year=current_academic_year, academic_term=term)
                for _assessment in assessments:
                    _assessment.title = new_title
                    assessments_to_update.append(_assessment)
                try:
                    Assessment.objects.bulk_update(assessments_to_update, ['title'])
                    return Response(status=200)
                except Exception:
                    return Response(status=400) 
            
        elif data['editType'] == 'description':
            with transaction.atomic():
                assessments_to_update = []
                new_description = data['newDescription']
                assessments = Assessment.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term)
                for _assessment in assessments:
                    _assessment.description = new_description
                    assessments_to_update.append(_assessment)
                try:
                    Assessment.objects.bulk_update(assessments_to_update, ['description'])
                    return Response(status=200)
                except Exception:
                    return Response(status=400) 
                
        elif data['editType'] == 'totalScore':
            existing_results = StudentResult.objects.filter(school=school, level=current_level, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term).exists()
            if existing_results:
                return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
            assessments_to_update = []
            new_total_score = float(data['newTotalScore'])
            if new_total_score <= 0:
                return Response({'message': "The total score of the assessment cannot be negative or zero(0)"}, status=400)
            assessments = Assessment.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term)
            new_total_score = Decimal(data['newTotalScore'])
            for _assessment in assessments:
                if _assessment.score and new_total_score < float(_assessment.score):
                    return Response({'message': "The total score of the assessment cannot be less than a student's score"}, status=400)
                _assessment.total_score = new_total_score
                assessments_to_update.append(_assessment)
            with transaction.atomic():
                try:
                    Assessment.objects.bulk_update(assessments_to_update, ['total_score'])
                    return Response(status=200)
                except Exception:
                    return Response(status=400) 
        
        elif data['editType'] == 'date':
            assessments_to_update = []
            new_date = data['newDate']
            new_date_object = datetime.strptime(new_date, '%Y-%m-%d').date()
            if (new_date_object > current_academic_year.end_date) or (new_date_object < current_academic_year.start_date):
                return Response({'message': f"The assessment date must be between the academic year start and end dates. That's {current_academic_year.start_date} and {current_academic_year.end_date}"}, status=400)
            assessments = Assessment.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term)
            for _assessment in assessments:
                _assessment.assessment_date = new_date
                assessments_to_update.append(_assessment)
            with transaction.atomic():
                try:
                    Assessment.objects.bulk_update(assessments_to_update, ['assessment_date'])
                    return Response(status=200)
                except Exception:
                    return Response({'message': 'An unexpected error occurred! try again later'}, status=400)
            
        elif data['editType'] == 'comment':
            new_comment = data['newComment']
            student = Student.objects.get(school=school, st_id=data['studentId'])
            with transaction.atomic():
                try:
                    assessment = Assessment.objects.get(school=school, level=current_level, teacher=teacher, subject=subject_obj, student=student, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term)
                    assessment.comment = new_comment
                    assessment.save()
                    return Response(status=200)
                except Exception:
                    return Response(status=400)     
            
        elif data['editType'] == 'score':
            existing_results = StudentResult.objects.filter(school=school, level=current_level, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term).exists()
            if existing_results:
                return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
            new_score = float(data['newScore'])
            if new_score < 0:
                return Response({'message': "The student's score cannot be negative"}, status=400)
            student = Student.objects.get(school=school, st_id=data['studentId'])
            with transaction.atomic():
                try:
                    new_score = Decimal(data['newScore'])
                    assessment = Assessment.objects.get(school=school, level=current_level, teacher=teacher, subject=subject_obj, student=student, student_class=students_class, title=assessment_title, academic_year=current_academic_year, academic_term=term)
                    if new_score > float(assessment.total_score):
                        return Response({"message": "The student's score cannot exceed the total assessment score"}, status=400)
                    assessment.score = new_score
                    assessment.save()
                    return Response(status=200)
                except Exception:
                    return Response(status=400)                
    
    elif data['type'] == 'deleteAssessment':
        students_class = Classe.objects.get(school=school, level=current_level, name=st_class_name)
        subject_obj = Subject.objects.get(name=subject_name)
        with transaction.atomic():
            try:
                existing_results = StudentResult.objects.filter(school=school, level=current_level, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term).exists()
                if existing_results:
                    return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
                assessments = Assessment.objects.filter(school=school, level=current_level, teacher=teacher, student_class=students_class, subject=subject_obj, title=assessment_title, academic_year=current_academic_year, academic_term=term)
                assessments.delete()
                return Response(status=200)
            except Exception:
                transaction.set_rollback(True)
                return Response(status=400)
        

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def teacher_exams(request):
    teacher = Staff.objects.select_related('school', 'current_role__level').get(user=request.user)
    school = teacher.school
    current_level = teacher.current_role.level
    data = request.data
    subject_name = data['subject']
    st_class_name = data['studentsClassName']
    year = data['year']
    current_academic_year = AcademicYear.objects.get(school=school, level=current_level, name=year)
    term = int(data['term'])
    if data['type'] == 'getFile':
        students = json.loads(data['selectedStudents'])
        if current_level.students_id:
            data = [
                ['STUDENT NAME', 'STUDENT ID', 'SCORE']
            ]
        else:
            data = [
                ['STUDENT NAME', 'STUDENT USERNAME', 'SCORE']
            ]

        if len(students) == 0:
            return Response({'message': f"You have already uploaded all the {subject_name} Exams scores for all students in this class"}, status=400)
        
        for _st in students:
            row = [_st['name'], _st['st_id'], '']
            data.append(row)

        wb = Workbook()
        ws = wb.active
        ws.title = subject_name

        ws.merge_cells('A1:I3')
        ws['A1'].value = f"SUBJECT: [{subject_name}]  ACADEMIC YEAR: [{year}]  CLASS: [{st_class_name}]  {current_academic_year.period_division} {term} EXAMS"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        count = 0
        container = ws['B5:D200']
        while count < len(data):
            for row_ws, row_data in zip(container, data):
                for cell, item in zip(row_ws, row_data):
                    cell.value = item

            count += 1

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15

        for row in ws['A3:F200']:
            for cell in row:
                cell.font = Font(size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws['B3:B200']:
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for cells in ws['A5:F200'][0]:
            cells.font = Font(bold=True, size=14)

        ws.protection.sheet = True
        for cell in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=4, max_col=4):
            cell[0].protection = Protection(locked=False)

        ws.protection.password = 'teamjn'

        filename = f"{subject_name.replace(' ', '-')}-{st_class_name}-Exams"
        byte_file = io.BytesIO()
        wb.save(byte_file)
        return send_file('excel', byte_file, filename)
    
    elif data['type'] == 'createExams':
        subject_obj = Subject.objects.get(name=subject_name)
        students_class = Classe.objects.get(school=school, level=current_level, name=st_class_name)
        existing_exams = Exam.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).exists()
        if existing_exams:
            return Response({'message': f"You have already created {subject_name} exams for this class"}, status=400)
        existing_results = StudentResult.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).exists()
        if existing_results:
            return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
        exams_total_score = data['totalScore']
        exams_obj = Exam.objects.create(
            school=school,
            level=current_level,
            teacher=teacher,
            student_class=students_class,
            subject=subject_obj,
            percentage=0,
            score=0,
            total_score=exams_total_score,
            academic_year=current_academic_year,
            academic_term=term
        )
        with transaction.atomic():
            try:
                exams_obj.save()
            except Exception:
                return Response(status=400)
            
        return Response({'message': "Operation Successful"})
        
    elif data['type'] == 'uploadWithFile':
        file = data['file']
        exams_total_score = float(data['totalScore'])
        try:
            wb = load_workbook(file)
        except Exception:
            return Response({'message': "Invalid file. Ensure you upload the excel file that you generated"}, status=400)
        
        ws = wb.active
        ws.protection = False
        if ws['A1'].value != f"SUBJECT: [{subject_name}]  ACADEMIC YEAR: [{year}]  CLASS: [{st_class_name}]  {current_academic_year.period_division} {term} EXAMS":
            return Response({'message': "Invalid file. Ensure you upload the excel file that you generated"}, status=400)
        
        cleaned_data = []
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=4):
            data = []
            for _cell in row:
                data.append(_cell.value)
            cleaned_data.append(data)
        
        valid_rows = [x for x in cleaned_data if x[0]]
        if len(valid_rows) == 0:
            return Response({'message': f"There are no students data in the uploaded file"}, status=400)
        
        students_to_create_ids = set()
        for row in valid_rows:
            score = row[2]
            try:
                float(score)
            except Exception:
                return Response({'message': f"The scores must be a number. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)
            if score > exams_total_score:
                return Response({'message': f"Scores cannot exceed the total score of the exams. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)
            elif score < 0:
                return Response({'message': f"Scores cannot be negative. Recheck the score for {row[0]}[{row[1]}] in the file"}, status=400)
            students_to_create_ids.add(row[1])
            
        students_class = Classe.objects.prefetch_related('students').get(school=school, level=current_level, name=st_class_name)
        students_to_create_objs = {x.st_id: x for x in students_class.students.filter(st_id__in=students_to_create_ids)}
        subject_obj = Subject.objects.get(schools=school, level=current_level, name=subject_name)
        exams_to_create = []
        for _student_data in valid_rows:
            st_id = _student_data[1]
            if st_id in students_to_create_ids:
                student = students_to_create_objs[st_id]
                st_exam = Exam(
                    school=school,
                    student=student,
                    subject=subject_obj,
                    teacher=teacher,
                    percentage=0,
                    total_score=exams_total_score,
                    student_class=students_class,
                    score=float(_student_data[2]),
                    academic_year=current_academic_year,
                    academic_term=term,
                    level=current_level,
                )
                exams_to_create.append(st_exam)
            else:
                return Response({'message': f"Invalid student {_student_data[0]}[{_student_data[1]}]. Make sure you don't change anything in the excel file except the scores"}, status=400)
        with transaction.atomic():
            try:
                Exam.objects.bulk_create(exams_to_create)
                exam_create_obj = Exam.objects.filter(school=school, level=current_level, teacher=teacher, student=None, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).first()
                if exam_create_obj:
                    exam_create_obj.delete()
            except IntegrityError:
                transaction.set_rollback(True)
                return Response({'message': f"You have already uploaded exams score for some of the students in the file. Click on get file to get an updated file"}, status=400)
            except Exception:
                transaction.set_rollback(True)
                return Response({'message': f"Invalid data! Ensure you don't delete or change anything in the excel file"}, status=400)
    
        all_exams = Exam.objects.select_related('student__user').filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, student__in=list(students_to_create_objs.values()), academic_year=current_academic_year, academic_term=term)
        all_exams_data = [{
            'name': f"{x.student.user.first_name} {x.student.user.last_name}",
            'st_id': x.student.st_id,
            'score': x.score,
        } for x in all_exams]
        
        return Response({'message': "The exams data has been uploaded successfully.", 'data': all_exams_data}, status=200)

    elif data['type'] == 'uploadWithoutFile':
        score = float(data['score'])
        exams_total_score = float(data['totalScore'])
        if score < 0:
            return Response({'message': "The score cannot be negative"}, status=400)
        if score > exams_total_score:
            return Response({'message': "The score must not exceed the total exams score"}, status=400)
        students_to_create_ids = json.loads(data['selectedStudents'])
        students_class = Classe.objects.prefetch_related(Prefetch('students', queryset=Student.objects.filter(st_id__in=students_to_create_ids))).get(school=school, level=current_level, name=st_class_name)
        subject_obj = Subject.objects.get(schools=school, level=current_level, name=subject_name)
        students_to_create_objs = {x.st_id: x for x in students_class.students.all()}
        exams_to_create = []
        for _st_id in students_to_create_ids:
            student = students_to_create_objs[_st_id]
            st_exam = Exam(
                school=school,
                student=student,
                subject=subject_obj,
                teacher=teacher,
                student_class=students_class,
                score=score,
                percentage=0,
                total_score=exams_total_score,
                academic_year=current_academic_year,
                academic_term=term,
                level=current_level,
            )
            exams_to_create.append(st_exam)
        with transaction.atomic():
            try:
                Exam.objects.bulk_create(exams_to_create)
                exam_create_obj = Exam.objects.filter(school=school, level=current_level, teacher=teacher, student=None, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=term).first()
                if exam_create_obj:
                    exam_create_obj.delete()
            except Exception:
                transaction.set_rollback(True)
                return Response(status=400)
        
        all_exams = Exam.objects.select_related('student__user').filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, student__in=list(students_to_create_objs.values()), academic_year=current_academic_year, academic_term=term)
        all_exams_data = [{
            'name': f"{x.student.user.first_name} {x.student.user.last_name}",
            'st_id': x.student.st_id,
            'score': x.score,
        } for x in all_exams]
        
        return Response({'message': "The exams data has been uploaded successfully.", 'data': all_exams_data}, status=200)

    elif data['type'] == 'editTotalScore':
        students_class = Classe.objects.get(school=school, level=current_level, name=st_class_name)
        subject_obj = Subject.objects.get(schools=school, level=current_level, name=subject_name)
        new_total_score = float(data['totalScore'])
        if new_total_score <= 0:
            return Response({'message': "The total score of the exams must be greater than zero(0)"}, status=400)
        if StudentResult.objects.filter(school=school, level=current_level, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term).exists():
            return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
        
        exams_to_update = []
        exams = Exam.objects.filter(school=school, level=current_level, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term)
        new_total_score = Decimal(data['totalScore'])
        for _exam in exams:
            if _exam.score > new_total_score:
                return Response({'message': "The total score of the exams must not be less than a students score"}, status=400)
            _exam.total_score = new_total_score
            exams_to_update.append(_exam)
            
        with transaction.atomic():
            try:
                Exam.objects.bulk_update(exams_to_update, ['total_score'])
            except Exception:
                transaction.set_rollback(True)
                return Response(status=400)
        return Response()
        
    
    elif data['type'] == 'editExamScore':
        new_score = Decimal(data['score'])
        with transaction.atomic():
            try:
                subject_obj = Subject.objects.get(schools=school, level=current_level, name=subject_name)
                student = Student.objects.select_related('st_class').get(school=school, level=current_level, st_id=data['studentId'])
                students_class = student.st_class
                if StudentResult.objects.filter(school=school, level=current_level, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term).exists():
                    return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
                st_exam = Exam.objects.get(school=school, level=current_level, student=student, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term)
                if new_score > st_exam.total_score:
                    return Response({'message': f"The student's score cannot exceed the total score for the exam"}, status=400)
                elif new_score < 0:
                    return Response({'message': f"The student's score cannot be negative"}, status=400)
                st_exam.score = new_score
                st_exam.save()
                
                return Response(status=200)
            except Exception:
                transaction.set_rollback(True)
                return Response(status=400)
            
    elif data['type'] == 'deleteExam':
        with transaction.atomic():
            try:
                students_class = Classe.objects.get(school=school, level=current_level, name=st_class_name)
                subject_obj = Subject.objects.get(schools=school, level=current_level, name=subject_name)
                existing_results = StudentResult.objects.filter(school=school, level=current_level, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term).exists()
                if existing_results:
                    return Response({'message': f"Results for {subject_name} in this class have already been generated. Please delete the existing results data and try again."}, status=400)
                exams = Exam.objects.filter(school=school, level=current_level, subject=subject_obj, student_class=students_class, teacher=teacher, academic_year=current_academic_year, academic_term=term)
                exams.delete()
            except Exception:
                transaction.set_rollback(True)
                return Response(status=400)
        
        return Response()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def teacher_students_results(request):
    teacher = Staff.objects.select_related('school', 'current_role__level').get(user=request.user)
    school = teacher.school
    current_level = teacher.current_role.level
    data = request.data
    current_academic_year = AcademicYear.objects.get(school=school, level=current_level, name=data['year'])
    current_term = int(data['term'])
    subject_obj = Subject.objects.get(name=data['subject'])
    students_class = Classe.objects.prefetch_related('students').get(school=school, level=current_level, name=data['studentsClassName'])
    
    if data['type'] == 'generateResults':
        result_data = json.loads(data['resultsData'])
        if (result_data['exams'] == 'yes' and (float(result_data['totalAssessmentPercentage']) + float(result_data['examsPercentage'])) != 100) or (result_data['exams'] == 'no' and float(result_data['totalAssessmentPercentage']) != 100):
            return Response({'message': "The total percentage must be 100"}, status=400)
        
        existing_results = StudentResult.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=current_term)
        if existing_results.exists():
            return Response({'message': f"{subject_obj.name} results for this class already exists"}, status=400)
        
        grading_system = GradingSystem.objects.prefetch_related(Prefetch('ranges', queryset=GradingSystemRange.objects.order_by('lower_limit'))).filter(schools=school, level=current_level).first()
        assessments_to_update = []
        exams_to_update = []
        results_to_create = []
        class_students = students_class.students.all()
        students_exams_data = {}
        students_assessment_objs = Assessment.objects.select_related('student').filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=current_term)
        students_assessments_title_data = defaultdict(dict)
        for _assessment in students_assessment_objs:
            students_assessments_title_data[_assessment.student.st_id][_assessment.title] = _assessment
        if result_data['exams'] == 'yes':
            students_exams_objs = Exam.objects.select_related('student').filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=current_term)
            for _exam in students_exams_objs:
                students_exams_data[_exam.student.st_id] = _exam
                    
        for _student in class_students:
            student_total_assessment_score = 0
            for _title_data in result_data['assessments']:
                student_assessment = students_assessments_title_data[_student.st_id][_title_data['title']]
                student_total_assessment_score += (float(student_assessment.score/student_assessment.total_score)*float(_title_data['percentage']))
                student_assessment.percentage = _title_data['percentage']
                assessments_to_update.append(student_assessment)
            
            student_exam_score = 0
            if result_data['exams'] == 'yes':
                student_exam = students_exams_data[_student.st_id]
                student_exam_score = float(student_exam.score/student_exam.total_score)*float(result_data['examsPercentage'])
                student_exam.percentage = result_data['examsPercentage']
                exams_to_update.append(student_exam)
                
            result = student_total_assessment_score + student_exam_score
            student_grade = next((item for item in grading_system.ranges.all() if item.lower_limit <= result and result <= (float(item.upper_limit) + 0.999)), None)
            if not student_grade:
                return Response({'message': f"Couldn't find a grade for a student with result of {result}"}, status=400)
            
            student_result = StudentResult(
                school=school,
                level=current_level,
                teacher=teacher,
                student=_student,
                subject=subject_obj,
                student_class=students_class,
                exam_percentage=result_data['examsPercentage'] if result_data['exams'] == 'yes' else 0,
                exam_score=student_exam_score,
                total_assessment_percentage=result_data['totalAssessmentPercentage'],
                total_assessment_score=student_total_assessment_score,
                result=result,
                position='',
                remark=student_grade.remark,
                grade=student_grade.label,
                academic_year = current_academic_year,
                academic_term=current_term,
            )
            results_to_create.append(student_result)
        
        p = inflect.engine()
        created_results = None
        with transaction.atomic():
            try:
                Assessment.objects.bulk_update(assessments_to_update, ['percentage'])
                Exam.objects.bulk_update(exams_to_update, ['percentage'])
                StudentResult.objects.bulk_create(results_to_create)
                created_results = StudentResult.objects.select_related('student').filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=current_term).order_by('-result')
                rank = 1
                previous_score = 0
                position_map = {}
                for index, _result in enumerate(created_results, start=1):
                    result = float(_result.result)
                    if result != previous_score:
                        rank = index
                    
                    previous_score = result
                    position_map[_result.id] = rank
                
                for _student_result in created_results:
                    _student_result.position = p.ordinal(position_map[_student_result.id])
                StudentResult.objects.bulk_update(created_results, ['position'])
                    
            except Exception:
                transaction.set_rollback(True)
                return Response(status=400)
            
        subject_data = {}
        first_result_obj = created_results.first()
        subject_data['total_assessment_percentage'] = first_result_obj.total_assessment_percentage
        subject_data['exam_percentage'] = first_result_obj.exam_percentage
        all_results = {}
        for _result in created_results:
            result_student = _result.student
            student_data = {
                'result': _result.result,
                'total_assessment_score': _result.total_assessment_score,
                'exam_score': _result.exam_score,
                'student': {'name': f"{_result.student.user.first_name} {_result.student.user.last_name}", 'st_id': result_student.st_id},
                'remark': _result.remark,
                'grade': _result.grade,
                'position': _result.position,
            }
            all_results[result_student.st_id] = student_data
        subject_data['student_results'] = all_results
        
        return Response(subject_data)
    
    elif data['type'] == 'deleteResults':
        with transaction.atomic():
            released_results = ReleasedResult.objects.filter(school=school, level=current_level, students_class=students_class, academic_year=current_academic_year, academic_term=current_term).exists()
            if released_results:
                return Response({'message': f"Results have already been released. Contact your school administrator if you still want to delete the students results data"}, status=400)
            
            results = StudentResult.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=current_term)
            results.delete()
            assessments = Assessment.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=current_term)
            if assessments.exists():
                assessments.update(percentage=0)
            exams = Exam.objects.filter(school=school, level=current_level, teacher=teacher, subject=subject_obj, student_class=students_class, academic_year=current_academic_year, academic_term=current_term)
            if exams.exists():
                exams.update(percentage=0)
        
        return Response()

